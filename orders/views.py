from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.contrib import messages
from django.db.models import Count, Sum, Q, Case, When, F, Value, IntegerField, CharField, Exists, OuterRef
from django.utils import timezone
import csv
import json
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Customer, Order, OrderItem, OrderSettings, PaymentMethod, ExtraProduct, OrderExtraItem, DeliveryBin, OrderUserMenuPermission, DeliveryCompletion
from django.contrib.auth import get_user_model
from .forms import CustomerForm, OrderForm, OrderItemFormSet, OrderSettingsForm, PaymentMethodForm, ExtraProductForm, OrderExtraItemFormSet, DeliveryBinForm
from sales.models import Product
from datetime import datetime, timedelta, date


def _get_extra_products_json():
    products = ExtraProduct.objects.filter(is_active=True)
    return json.dumps([
        {'id': p.pk, 'name': p.name, 'unit_price': int(p.unit_price)}
        for p in products
    ], ensure_ascii=False)


@login_required
def order_list(request):
    return redirect('orders:regular_dashboard')


@login_required
def order_create(request, customer_id=None):
    today = timezone.localdate()
    initial = {'order_date': today, 'delivery_date': today}
    delivery_date_str = request.GET.get('delivery_date')
    if delivery_date_str:
        try:
            initial['delivery_date'] = datetime.strptime(delivery_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if customer_id:
        customer = get_object_or_404(Customer, pk=customer_id, is_active=True)
        initial['customer'] = customer
        if customer.notes:
            initial['notes'] = customer.notes

    if request.method == 'POST':
        form = OrderForm(request.POST)
        formset = OrderItemFormSet(request.POST, prefix='items')
        extra_formset = OrderExtraItemFormSet(request.POST, prefix='extra_items')
        if form.is_valid() and formset.is_valid() and extra_formset.is_valid():
            order = form.save(commit=False)
            order.created_by = request.user
            order.save()
            formset.instance = order
            items = formset.save(commit=False)
            for item in items:
                total_qty = item.quantity_large + item.quantity_regular + item.quantity_small
                if total_qty == 0:
                    continue
                if item.product and not item.product_name:
                    item.product_name = item.product.name
                item.quantity = total_qty
                item.save()
            for obj in formset.deleted_objects:
                obj.delete()
            extra_formset.instance = order
            extra_items = extra_formset.save(commit=False)
            for ei in extra_items:
                if not ei.product_name and not ei.extra_product_id:
                    continue
                if not ei.quantity or not ei.unit_price:
                    continue
                if not ei.product_name and ei.extra_product:
                    ei.product_name = ei.extra_product.name
                ei.subtotal = ei.unit_price * ei.quantity
                ei.save()
            for obj in extra_formset.deleted_objects:
                obj.delete()
            messages.success(request, f'受注 {order.order_number} を登録しました。')
            return redirect('orders:order_detail', pk=order.pk)
    else:
        form = OrderForm(initial=initial)
        formset = OrderItemFormSet(prefix='items')
        extra_formset = OrderExtraItemFormSet(prefix='extra_items')

    one_month_ago = timezone.localdate() - timedelta(days=30)
    customers = (
        Customer.objects.filter(is_active=True)
        .annotate(
            has_recent_order=Exists(
                Order.objects.filter(customer=OuterRef('pk'), order_date__gte=one_month_ago)
            )
        )
        .annotate(
            group_order=Case(
                When(is_regular=True, then=Value(0)),
                When(has_recent_order=True, then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            ),
            group_label=Case(
                When(is_regular=True, then=Value('定期顧客')),
                When(has_recent_order=True, then=Value('直近1ヶ月')),
                default=Value('その他'),
                output_field=CharField(),
            ),
            sort_name=Case(
                When(customer_type='B2B', then=F('company_name')),
                default=F('name'),
                output_field=CharField(),
            ),
        )
        .order_by('group_order', 'sort_name')
    )
    is_catering = False
    if customer_id:
        _c = get_object_or_404(Customer, pk=customer_id, is_active=True)
        is_catering = _c.bento_type == 'CATERING'
    context = {
        'form': form,
        'formset': formset,
        'extra_formset': extra_formset,
        'extra_products_json': _get_extra_products_json(),
        'customers': customers,
        'is_edit': False,
        'is_catering': is_catering,
    }
    return render(request, 'orders/order_form.html', context)


@login_required
def order_edit(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        form = OrderForm(request.POST, instance=order)
        formset = OrderItemFormSet(request.POST, instance=order, prefix='items')
        extra_formset = OrderExtraItemFormSet(request.POST, instance=order, prefix='extra_items')
        if form.is_valid() and formset.is_valid() and extra_formset.is_valid():
            form.save()
            items = formset.save(commit=False)
            for item in items:
                total_qty = item.quantity_large + item.quantity_regular + item.quantity_small
                if total_qty == 0:
                    if item.pk:
                        item.delete()
                    continue
                if item.product and not item.product_name:
                    item.product_name = item.product.name
                item.quantity = total_qty
                item.save()
            for obj in formset.deleted_objects:
                obj.delete()
            extra_items = extra_formset.save(commit=False)
            for ei in extra_items:
                if not ei.product_name and not ei.extra_product_id:
                    continue
                if not ei.quantity or not ei.unit_price:
                    continue
                if not ei.product_name and ei.extra_product:
                    ei.product_name = ei.extra_product.name
                ei.subtotal = ei.unit_price * ei.quantity
                ei.save()
            for obj in extra_formset.deleted_objects:
                obj.delete()
            messages.success(request, f'受注 {order.order_number} を更新しました。')
            return redirect('orders:order_detail', pk=order.pk)
    else:
        form = OrderForm(instance=order)
        formset = OrderItemFormSet(instance=order, prefix='items')
        extra_formset = OrderExtraItemFormSet(instance=order, prefix='extra_items')

    one_month_ago = timezone.localdate() - timedelta(days=30)
    customers = (
        Customer.objects.filter(is_active=True)
        .annotate(
            has_recent_order=Exists(
                Order.objects.filter(customer=OuterRef('pk'), order_date__gte=one_month_ago)
            )
        )
        .annotate(
            group_order=Case(
                When(is_regular=True, then=Value(0)),
                When(has_recent_order=True, then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            ),
            group_label=Case(
                When(is_regular=True, then=Value('定期顧客')),
                When(has_recent_order=True, then=Value('直近1ヶ月')),
                default=Value('その他'),
                output_field=CharField(),
            ),
            sort_name=Case(
                When(customer_type='B2B', then=F('company_name')),
                default=F('name'),
                output_field=CharField(),
            ),
        )
        .order_by('group_order', 'sort_name')
    )
    context = {
        'form': form,
        'formset': formset,
        'extra_formset': extra_formset,
        'extra_products_json': _get_extra_products_json(),
        'order': order,
        'customers': customers,
        'is_edit': True,
        'is_catering': order.customer.bento_type == 'CATERING',
    }
    return render(request, 'orders/order_form.html', context)


@login_required
def order_detail(request, pk):
    order = get_object_or_404(
        Order.objects.select_related('customer').prefetch_related('items', 'extra_items'),
        pk=pk
    )
    return render(request, 'orders/order_detail.html', {'order': order})


@login_required
def order_delete(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        order_number = order.order_number
        order.delete()
        messages.success(request, f'受注 {order_number} を削除しました。')
        return redirect('orders:regular_dashboard')
    return render(request, 'orders/order_confirm_delete.html', {'order': order})


@login_required
def order_pdf(request, pk):
    order = get_object_or_404(
        Order.objects.select_related('customer').prefetch_related('items', 'extra_items'),
        pk=pk
    )
    from .pdf import generate_delivery_slip
    buffer = generate_delivery_slip(order)
    if not order.pdf_printed_at:
        order.pdf_printed_at = timezone.now()
        order.save(update_fields=['pdf_printed_at'])
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="delivery_slip_{order.order_number}.pdf"'
    return response


@login_required
def order_print_preview(request, pk):
    order = get_object_or_404(
        Order.objects.select_related('customer').prefetch_related('items', 'extra_items'),
        pk=pk
    )
    from .pdf import build_delivery_slip_context
    context = build_delivery_slip_context(order)
    if not order.pdf_printed_at:
        order.pdf_printed_at = timezone.now()
        order.save(update_fields=['pdf_printed_at'])
    return render(request, 'orders/delivery_slip_preview.html', context)


@login_required
def receipt_print_preview(request, pk):
    order = get_object_or_404(Order.objects.select_related('customer'), pk=pk)
    from .pdf import build_receipt_context
    context = build_receipt_context(order)
    return render(request, 'orders/receipt_preview.html', context)


@login_required
def update_receipt_memo(request, pk):
    if request.method != 'POST':
        return HttpResponseBadRequest()
    order = get_object_or_404(Order, pk=pk)
    memo = request.POST.get('receipt_memo', 'お弁当代').strip() or 'お弁当代'
    Order.objects.filter(pk=pk).update(receipt_memo=memo)
    return redirect(reverse('orders:receipt_print', kwargs={'pk': pk}) + '?updated=1')


@login_required
def receipt_pdf(request, pk):
    order = get_object_or_404(Order.objects.select_related('customer'), pk=pk)
    from .pdf import generate_receipt
    buffer = generate_receipt(order)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="receipt_{order.order_number}.pdf"'
    return response


@login_required
def batch_delivery_slip_pdf(request):
    if request.method != 'POST':
        return HttpResponseBadRequest('POST required')
    pks = [int(v) for v in request.POST.getlist('order_pks') if v.strip().isdigit()]
    if not pks:
        return HttpResponseBadRequest('注文が選択されていません')
    orders = (Order.objects
              .filter(pk__in=pks)
              .select_related('customer')
              .prefetch_related('items', 'extra_items')
              .order_by('customer__company_name', 'customer__name'))
    from .pdf import build_delivery_slip_context
    contexts = [build_delivery_slip_context(order) for order in orders]
    needs_zoom = any(ctx.get('customer_notes') and ctx.get('has_extra_items') for ctx in contexts)
    orders.filter(pdf_printed_at__isnull=True).update(pdf_printed_at=timezone.now())
    return render(request, 'orders/batch_delivery_slip_preview.html', {
        'orders_contexts': contexts,
        'needs_zoom': needs_zoom,
    })


@login_required
def batch_receipt_pdf(request):
    if request.method != 'POST':
        return HttpResponseBadRequest('POST required')
    pks = [int(v) for v in request.POST.getlist('order_pks') if v.strip().isdigit()]
    if not pks:
        return HttpResponseBadRequest('注文が選択されていません')
    orders = (Order.objects
              .filter(pk__in=pks)
              .select_related('customer')
              .order_by('customer__company_name', 'customer__name'))
    from .pdf import build_receipt_context
    contexts = [build_receipt_context(order) for order in orders]
    return render(request, 'orders/batch_receipt_preview.html', {'orders_contexts': contexts})


# --- Customer views ---

@login_required
def customer_list(request):
    customers = Customer.objects.filter(is_active=True)
    q = request.GET.get('q', '').strip()
    customer_type = request.GET.get('type', '')
    if q:
        from django.db.models import Q
        customers = customers.filter(
            Q(name__icontains=q) | Q(company_name__icontains=q) |
            Q(phone__icontains=q) | Q(contact_person__icontains=q)
        )
    if customer_type:
        customers = customers.filter(customer_type=customer_type)
    context = {
        'customers': customers,
        'q': q,
        'customer_type': customer_type,
    }
    return render(request, 'orders/customer_list.html', context)


@login_required
def customer_create(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save()
            messages.success(request, f'顧客「{customer}」を登録しました。')
            return redirect('orders:customer_list')
    else:
        form = CustomerForm()
    return render(request, 'orders/customer_form.html', {
        'form': form, 'is_edit': False,
        'day_fields': [('月', 'schedule_mon'), ('火', 'schedule_tue'), ('水', 'schedule_wed'),
                       ('木', 'schedule_thu'), ('金', 'schedule_fri'), ('土', 'schedule_sat'), ('日', 'schedule_sun')],
    })


@login_required
def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    recent_orders = customer.orders.prefetch_related('items')[:20]
    return render(request, 'orders/customer_detail.html', {
        'customer': customer,
        'recent_orders': recent_orders,
    })


@login_required
def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, f'顧客「{customer}」を更新しました。')
            return redirect('orders:customer_detail', pk=customer.pk)
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'orders/customer_form.html', {
        'form': form, 'customer': customer, 'is_edit': True,
        'day_fields': [('月', 'schedule_mon'), ('火', 'schedule_tue'), ('水', 'schedule_wed'),
                       ('木', 'schedule_thu'), ('金', 'schedule_fri'), ('土', 'schedule_sat'), ('日', 'schedule_sun')],
    })


@login_required
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer.is_active = False
        customer.save()
        messages.success(request, f'顧客「{customer}」を無効にしました。')
        return redirect('orders:customer_list')
    return render(request, 'orders/customer_confirm_delete.html', {'customer': customer})


# --- API endpoints ---

@login_required
def api_customer_info(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    return JsonResponse({
        'price_type': customer.price_type,
        'name': str(customer),
        'customer_type': customer.customer_type,
        'bento_type': customer.bento_type,
        'notes': customer.notes,
    })


@login_required
def api_products(request, date):
    try:
        if '-' in date:
            target = datetime.strptime(date, '%Y-%m-%d').date()
        else:
            target = datetime.strptime(date, '%Y%m%d').date()
    except ValueError:
        return JsonResponse({'products': []})

    candidates = [(target - timedelta(days=i)).strftime('%Y-%m-%d') for i in range(7)]
    products = Product.objects.filter(week__in=candidates).order_by('no')
    data = [
        {
            'id': p.id,
            'no': p.no,
            'name': p.name,
            'price_A': int(p.price_A),
            'price_B': int(p.price_B),
            'price_C': int(p.price_C),
            'container_type': p.container_type,
        }
        for p in products
    ]
    return JsonResponse({'products': data})


@login_required
def regular_order_dashboard(request):
    today = timezone.localdate()
    target_date_str = request.GET.get('date')
    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            target_date = today
    else:
        target_date = today

    weekday = target_date.weekday()  # 0=月, 1=火, 2=水, 3=木, 4=金, 5=土, 6=日
    DAY_FIELDS = ['schedule_mon', 'schedule_tue', 'schedule_wed',
                  'schedule_thu', 'schedule_fri', 'schedule_sat', 'schedule_sun']
    WEEKDAY_LABELS = ['月曜日', '火曜日', '水曜日', '木曜日', '金曜日', '土曜日', '日曜日']
    weekday_label = WEEKDAY_LABELS[weekday]

    # 平日毎日の顧客（月〜金のみ対象）
    if weekday < 5:
        weekday_customers = Customer.objects.filter(
            is_active=True, is_regular=True, regular_type='WEEKDAY'
        )
    else:
        weekday_customers = Customer.objects.none()

    # 曜日指定の顧客
    day_field = DAY_FIELDS[weekday]
    custom_customers = Customer.objects.filter(
        is_active=True, is_regular=True, regular_type='CUSTOM',
        **{day_field: True}
    )

    regular_customers = (weekday_customers | custom_customers).select_related('payment_method')

    orders_for_date = Order.objects.filter(
        delivery_date=target_date,
        customer__in=regular_customers
    ).select_related('customer').prefetch_related('items')

    order_map = {o.customer_id: o for o in orders_for_date}

    dashboard_rows = []
    for customer in regular_customers:
        order = order_map.get(customer.pk)
        dashboard_rows.append({
            'customer': customer,
            'order': order,
            'has_order': order is not None,
            'pdf_printed': order.pdf_printed_at is not None if order else False,
            'total_quantity': order.total_quantity if order else 0,
            'total': order.total if order else 0,
        })

    total_customers = len(dashboard_rows)
    ordered_count = sum(1 for r in dashboard_rows if r['has_order'])

    # その他の注文（非定期顧客）
    other_orders = Order.objects.filter(
        delivery_date=target_date
    ).exclude(
        customer__in=regular_customers
    ).select_related('customer', 'customer__payment_method').prefetch_related('items')

    # PDF印刷済み数: その日の全注文から算出
    all_orders_for_date = Order.objects.filter(delivery_date=target_date)
    all_orders_count = all_orders_for_date.count()
    all_printed_count = all_orders_for_date.filter(pdf_printed_at__isnull=False).count()

    # メニュー別注文数集計（No1〜10すべて表示）
    week_candidates = [(target_date - timedelta(days=i)).strftime('%Y-%m-%d') for i in range(7)]
    week_products = Product.objects.filter(week__in=week_candidates, no__lte=10).order_by('no')

    order_totals = {
        row['product__id']: row
        for row in OrderItem.objects.filter(
            order__delivery_date=target_date
        ).values('product__id').annotate(
            total_large=Sum('quantity_large'),
            total_regular=Sum('quantity_regular'),
            total_small=Sum('quantity_small'),
            total_quantity=Sum('quantity'),
        )
    }

    menu_summary = []
    for p in week_products:
        totals = order_totals.get(p.id, {})
        menu_summary.append({
            'no': p.no,
            'product_name': p.name,
            'total_large': totals.get('total_large', 0) or 0,
            'total_regular': totals.get('total_regular', 0) or 0,
            'total_small': totals.get('total_small', 0) or 0,
            'total_quantity': totals.get('total_quantity', 0) or 0,
        })

    menu_totals = {
        'large': sum(m['total_large'] for m in menu_summary),
        'regular': sum(m['total_regular'] for m in menu_summary),
        'small': sum(m['total_small'] for m in menu_summary),
        'quantity': sum(m['total_quantity'] for m in menu_summary),
    }

    # 支払い方法別合計集計
    _pm_map = {}
    for row in dashboard_rows:
        if row['has_order']:
            pm = row['customer'].payment_method
            _key = pm.id if pm else 0
            if _key not in _pm_map:
                _pm_map[_key] = {
                    'name': pm.name if pm else 'その他',
                    'total': 0,
                    'count': 0,
                    'sort_order': pm.sort_order if pm else 999,
                }
            _pm_map[_key]['total'] += row['total']
            _pm_map[_key]['count'] += 1
    for order in other_orders:
        pm = order.customer.payment_method
        _key = pm.id if pm else 0
        if _key not in _pm_map:
            _pm_map[_key] = {
                'name': pm.name if pm else 'その他',
                'total': 0,
                'count': 0,
                'sort_order': pm.sort_order if pm else 999,
            }
        _pm_map[_key]['total'] += order.total
        _pm_map[_key]['count'] += 1
    payment_method_totals = sorted(_pm_map.values(), key=lambda x: x['sort_order'])

    context = {
        'target_date': target_date,
        'target_date_str': target_date.strftime('%Y-%m-%d'),
        'dashboard_rows': dashboard_rows,
        'total_customers': total_customers,
        'ordered_count': ordered_count,
        'not_ordered_count': total_customers - ordered_count,
        'other_orders': other_orders,
        'all_orders_count': all_orders_count,
        'all_printed_count': all_printed_count,
        'menu_summary': menu_summary,
        'menu_totals': menu_totals,
        'today_display': today.strftime('%Y年%m月%d日'),
        'payment_method_totals': payment_method_totals,
        'weekday_label': weekday_label,
    }
    return render(request, 'orders/regular_dashboard.html', context)


@login_required
def api_check_duplicate(request):
    customer_id = request.GET.get('customer_id')
    delivery_date = request.GET.get('delivery_date')
    exclude_pk = request.GET.get('exclude_pk')
    if not customer_id or not delivery_date:
        return JsonResponse({'duplicate': False})
    qs = Order.objects.filter(customer_id=customer_id, delivery_date=delivery_date)
    if exclude_pk:
        qs = qs.exclude(pk=exclude_pk)
    return JsonResponse({'duplicate': qs.exists()})


@login_required
def api_customers_with_orders(request, date):
    """指定納品日に注文済みの顧客IDリストを返す"""
    try:
        target_date = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'customer_ids': []})
    exclude_pk = request.GET.get('exclude_pk')
    qs = Order.objects.filter(delivery_date=target_date)
    if exclude_pk:
        qs = qs.exclude(pk=exclude_pk)
    customer_ids = list(qs.values_list('customer_id', flat=True))
    return JsonResponse({'customer_ids': customer_ids})


@login_required
def api_order_totals(request, date):
    """納品日ごとのメニュー別注文数合計を返すAPI"""
    try:
        target_date = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'totals': []})

    qs = OrderItem.objects.filter(
        order__delivery_date=target_date,
        product__isnull=False,
        order__customer__include_in_total=True,
    )

    exclude_order_pk = request.GET.get('exclude_order_pk')
    if exclude_order_pk:
        qs = qs.exclude(order__pk=exclude_order_pk)

    order_totals = {
        row['product__id']: row
        for row in qs.values('product__id').annotate(
            total_large=Sum('quantity_large'),
            total_regular=Sum('quantity_regular'),
            total_small=Sum('quantity_small'),
            total_quantity=Sum('quantity'),
        )
    }

    totals = [
        {
            'product_id': pid,
            'total_large': vals.get('total_large', 0) or 0,
            'total_regular': vals.get('total_regular', 0) or 0,
            'total_small': vals.get('total_small', 0) or 0,
            'total_quantity': vals.get('total_quantity', 0) or 0,
        }
        for pid, vals in order_totals.items()
    ]

    return JsonResponse({'totals': totals})


# --- PaymentMethod views ---

@login_required
def payment_method_list(request):
    payment_methods = PaymentMethod.objects.all()
    return render(request, 'orders/payment_method_list.html', {'payment_methods': payment_methods})


@login_required
def payment_method_create(request):
    if request.method == 'POST':
        form = PaymentMethodForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'支払い方法「{form.instance.name}」を登録しました。')
            return redirect('orders:order_settings')
    else:
        form = PaymentMethodForm()
    return render(request, 'orders/payment_method_form.html', {'form': form, 'is_edit': False})


@login_required
def payment_method_edit(request, pk):
    payment_method = get_object_or_404(PaymentMethod, pk=pk)
    if request.method == 'POST':
        form = PaymentMethodForm(request.POST, instance=payment_method)
        if form.is_valid():
            form.save()
            messages.success(request, f'支払い方法「{payment_method.name}」を更新しました。')
            return redirect('orders:order_settings')
    else:
        form = PaymentMethodForm(instance=payment_method)
    return render(request, 'orders/payment_method_form.html', {'form': form, 'payment_method': payment_method, 'is_edit': True})


@login_required
def payment_method_delete(request, pk):
    payment_method = get_object_or_404(PaymentMethod, pk=pk)
    if request.method == 'POST':
        payment_method.is_active = False
        payment_method.save()
        messages.success(request, f'支払い方法「{payment_method.name}」を無効にしました。')
        return redirect('orders:order_settings')
    return render(request, 'orders/payment_method_confirm_delete.html', {'payment_method': payment_method})


# --- Settings view ---

@login_required
def order_settings(request):
    User = get_user_model()
    settings_obj = OrderSettings.load()
    payment_methods = PaymentMethod.objects.all()
    extra_products = ExtraProduct.objects.all()
    delivery_bins = DeliveryBin.objects.all()

    orders_users = User.objects.filter(
        menu_permission__can_view_orders=True,
        is_active=True
    ).order_by('last_name', 'first_name')

    if request.method == 'POST':
        if request.POST.get('action') == 'menu_permissions':
            for u in orders_users:
                perm, _ = OrderUserMenuPermission.objects.get_or_create(user=u)
                perm.can_view_dashboard = f'perm_{u.id}_dashboard' in request.POST
                perm.can_view_delivery_list = f'perm_{u.id}_delivery_list' in request.POST
                perm.can_view_new_order = f'perm_{u.id}_new_order' in request.POST
                perm.can_view_customers = f'perm_{u.id}_customers' in request.POST
                perm.can_view_settings = f'perm_{u.id}_settings' in request.POST
                perm.can_view_csv_export = f'perm_{u.id}_csv_export' in request.POST
                perm.save()
            messages.success(request, 'ユーザーメニュー設定を保存しました。')
            return redirect('orders:order_settings')
        else:
            form = OrderSettingsForm(request.POST, instance=settings_obj)
            if form.is_valid():
                form.save()
                messages.success(request, '設定を保存しました。')
                return redirect('orders:order_settings')
    else:
        form = OrderSettingsForm(instance=settings_obj)

    user_perms = []
    for u in orders_users:
        perm, _ = OrderUserMenuPermission.objects.get_or_create(user=u)
        user_perms.append({'user': u, 'perm': perm})

    context = {
        'form': form,
        'payment_methods': payment_methods,
        'extra_products': extra_products,
        'delivery_bins': delivery_bins,
        'user_perms': user_perms,
    }
    return render(request, 'orders/order_settings.html', context)


# --- ExtraProduct views ---

@login_required
def extra_product_create(request):
    if request.method == 'POST':
        form = ExtraProductForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'追加商品「{form.instance.name}」を登録しました。')
            return redirect('orders:order_settings')
    else:
        form = ExtraProductForm()
    return render(request, 'orders/extra_product_form.html', {'form': form, 'is_edit': False})


@login_required
def extra_product_edit(request, pk):
    extra_product = get_object_or_404(ExtraProduct, pk=pk)
    if request.method == 'POST':
        form = ExtraProductForm(request.POST, instance=extra_product)
        if form.is_valid():
            form.save()
            messages.success(request, f'追加商品「{extra_product.name}」を更新しました。')
            return redirect('orders:order_settings')
    else:
        form = ExtraProductForm(instance=extra_product)
    return render(request, 'orders/extra_product_form.html', {
        'form': form, 'extra_product': extra_product, 'is_edit': True
    })


@login_required
def extra_product_delete(request, pk):
    extra_product = get_object_or_404(ExtraProduct, pk=pk)
    if request.method == 'POST':
        extra_product.is_active = False
        extra_product.save()
        messages.success(request, f'追加商品「{extra_product.name}」を無効にしました。')
        return redirect('orders:order_settings')
    return render(request, 'orders/extra_product_confirm_delete.html', {'extra_product': extra_product})


@login_required
def api_extra_products(request):
    products = ExtraProduct.objects.filter(is_active=True)
    data = [{'id': p.pk, 'name': p.name, 'unit_price': int(p.unit_price)} for p in products]
    return JsonResponse({'extra_products': data})


# --- DeliveryBin views ---

@login_required
def delivery_bin_create(request):
    if request.method == 'POST':
        form = DeliveryBinForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'配達便「{form.instance.name}」を登録しました。')
            return redirect('orders:order_settings')
    else:
        form = DeliveryBinForm()
    return render(request, 'orders/delivery_bin_form.html', {'form': form, 'is_edit': False})


@login_required
def delivery_bin_edit(request, pk):
    delivery_bin = get_object_or_404(DeliveryBin, pk=pk)
    if request.method == 'POST':
        form = DeliveryBinForm(request.POST, instance=delivery_bin)
        if form.is_valid():
            form.save()
            messages.success(request, f'配達便「{delivery_bin.name}」を更新しました。')
            return redirect('orders:order_settings')
    else:
        form = DeliveryBinForm(instance=delivery_bin)
    return render(request, 'orders/delivery_bin_form.html', {
        'form': form, 'delivery_bin': delivery_bin, 'is_edit': True
    })


@login_required
def delivery_bin_delete(request, pk):
    delivery_bin = get_object_or_404(DeliveryBin, pk=pk)
    if request.method == 'POST':
        delivery_bin.is_active = False
        delivery_bin.save()
        messages.success(request, f'配達便「{delivery_bin.name}」を無効にしました。')
        return redirect('orders:order_settings')
    return render(request, 'orders/delivery_bin_confirm_delete.html', {'delivery_bin': delivery_bin})


# --- Delivery List views ---

def _get_scheduled_customer_ids(target_date):
    """指定日にスケジュールされているレギュラー顧客のIDセットを返す"""
    weekday = target_date.weekday()
    DAY_FIELDS = ['schedule_mon', 'schedule_tue', 'schedule_wed',
                  'schedule_thu', 'schedule_fri', 'schedule_sat', 'schedule_sun']
    if weekday < 5:
        weekday_qs = Customer.objects.filter(
            is_active=True, is_regular=True, regular_type='WEEKDAY'
        )
    else:
        weekday_qs = Customer.objects.none()
    custom_qs = Customer.objects.filter(
        is_active=True, is_regular=True, regular_type='CUSTOM',
        **{DAY_FIELDS[weekday]: True}
    )
    return set((weekday_qs | custom_qs).values_list('pk', flat=True))


def _parse_target_date(request):
    today = timezone.localdate()
    date_str = request.GET.get('date')
    if date_str:
        try:
            return datetime.strptime(date_str, '%Y-%m-%d').date(), today
        except ValueError:
            pass
    return today, today


@login_required
def order_export_page(request):
    """CSV出力の専用ページ（日付範囲フォーム表示）"""
    today = timezone.localdate()
    default_from = today.replace(day=1)
    date_from = request.GET.get('date_from', default_from.strftime('%Y-%m-%d'))
    date_to   = request.GET.get('date_to',   today.strftime('%Y-%m-%d'))
    return render(request, 'orders/order_export.html', {
        'date_from': date_from,
        'date_to': date_to,
    })


@login_required
def order_csv_export(request):
    """売上データExcelファイルをダウンロードする（シート1: 売上データ、シート2: 顧客別価格表）"""
    today = timezone.localdate()
    date_from_str = request.GET.get('date_from')
    date_to_str   = request.GET.get('date_to')

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else today.replace(day=1)
    except ValueError:
        date_from = today.replace(day=1)

    try:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else today
    except ValueError:
        date_to = today

    orders = (
        Order.objects
        .filter(delivery_date__gte=date_from, delivery_date__lte=date_to)
        .select_related('customer', 'customer__payment_method', 'customer__delivery_bin')
        .prefetch_related('items', 'extra_items')
        .order_by('delivery_date', 'order_number')
    )

    wb = openpyxl.Workbook()

    # ---- シート1: 売上データ ----
    ws1 = wb.active
    ws1.title = '売上データ'

    header_fill = PatternFill(fill_type='solid', fgColor='4CAF50')
    header_font = Font(bold=True, color='FFFFFF')

    headers1 = [
        'ORID', '日付', '顧客種別', '顧客名', '部署名', '配達便', 'お弁当種別',
        '支払い方法', '価格タイプ',
        '★お弁当注文数', 'お弁当注文数', '大盛りご飯数', 'お弁当合計金額',
        '追加商品注文数', '追加商品合計金額', '総計金額',
    ]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for order in orders:
        c = order.customer
        customer_name = c.company_name if c.customer_type == 'B2B' else c.name
        delivery_bin  = c.delivery_bin.name if c.delivery_bin else ''
        payment_name  = c.payment_method.name if c.payment_method else ''

        items = list(order.items.all())
        bento_qty_star   = sum(item.quantity for item in items if '★' in item.product_name)
        bento_qty_normal = sum(item.quantity for item in items if '★' not in item.product_name)
        bento_qty_large  = sum(item.quantity_large for item in items)
        bento_total = sum(item.subtotal for item in items)
        extra_qty   = sum(ei.quantity for ei in order.extra_items.all())
        extra_total = sum(ei.subtotal for ei in order.extra_items.all())
        grand_total = bento_total + extra_total

        ws1.append([
            order.order_number,
            order.delivery_date.strftime('%Y-%m-%d'),
            c.customer_type,
            customer_name,
            c.department,
            delivery_bin,
            c.bento_type,
            payment_name,
            c.price_type,
            bento_qty_star,
            bento_qty_normal,
            bento_qty_large,
            int(bento_total),
            extra_qty,
            int(extra_total),
            int(grand_total),
        ])

    # 列幅の自動調整（シート1）
    for col in ws1.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ---- シート2: 顧客別価格表 ----
    ws2 = wb.create_sheet(title='顧客別価格表')

    # 期間内の最新週のProductから A/B/C 価格を取得
    # product_prices[価格タイプ]['star'|'normal'|'large'] = 単価（円）
    product_prices = {
        'A': {'star': None, 'normal': None, 'large': None},
        'B': {'star': None, 'normal': None, 'large': None},
        'C': {'star': None, 'normal': None, 'large': None},
    }
    latest_week = (
        Product.objects
        .filter(week__lte=date_to.strftime('%Y%m%d'))
        .order_by('-week')
        .values_list('week', flat=True)
        .first()
    )
    if latest_week:
        for prod in Product.objects.filter(week=latest_week):
            if '大盛りごはん' in prod.name:
                key = 'large'
            elif '★' in prod.name:
                key = 'star'
            else:
                key = 'normal'
            for pt in ('A', 'B', 'C'):
                if product_prices[pt][key] is None:
                    val = getattr(prod, f'price_{pt}')
                    product_prices[pt][key] = int(val) if val else None

    week_label = f"（{latest_week[:4]}/{latest_week[4:6]}/{latest_week[6:]}週）" if latest_week else ''

    header_fill2 = PatternFill(fill_type='solid', fgColor='1565C0')
    headers2 = [
        '顧客名', '顧客種別', '配達便', 'お弁当種別', '価格タイプ',
        f'★あり単価{week_label}', f'★なし単価{week_label}', '大盛り割増',
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill2
        cell.alignment = Alignment(horizontal='center')

    customers = (
        Customer.objects
        .filter(is_active=True)
        .select_related('delivery_bin')
        .order_by('delivery_bin__name', 'company_name', 'name')
    )

    custom_fill = PatternFill(fill_type='solid', fgColor='FFF9C4')  # 独自価格行を薄黄色でハイライト

    for c in customers:
        customer_name = c.company_name if c.customer_type == 'B2B' else c.name
        delivery_bin  = c.delivery_bin.name if c.delivery_bin else ''
        price_type_label = dict(Customer.PRICE_TYPE_CHOICES).get(c.price_type, c.price_type)
        customer_type_label = dict(Customer.CUSTOMER_TYPE_CHOICES).get(c.customer_type, c.customer_type)
        bento_type_label = dict(Customer.BENTO_TYPE_CHOICES).get(c.bento_type, c.bento_type)

        if c.price_type == 'CUSTOM':
            price_star   = int(c.custom_price_star)   if c.custom_price_star   is not None else ''
            price_normal = int(c.custom_price_normal) if c.custom_price_normal is not None else ''
            price_large  = int(c.custom_price_large)  if c.custom_price_large  is not None else ''
        else:
            pt = c.price_type  # 'A', 'B', or 'C'
            price_star   = product_prices[pt]['star']   if pt in product_prices else ''
            price_normal = product_prices[pt]['normal'] if pt in product_prices else ''
            price_large  = product_prices[pt]['large']  if pt in product_prices else ''

        row = ws2.max_row + 1
        ws2.append([
            customer_name,
            customer_type_label,
            delivery_bin,
            bento_type_label,
            price_type_label,
            price_star if price_star is not None else '',
            price_normal if price_normal is not None else '',
            price_large,
        ])
        if c.price_type == 'CUSTOM':
            for cell in ws2[row]:
                cell.fill = custom_fill

    # 列幅の自動調整（シート2）
    for col in ws2.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # レスポンス出力
    filename = f"orders_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def delivery_list(request):
    """配達便一覧インデックスページ（管理者向け）"""
    target_date, today = _parse_target_date(request)
    weekday = target_date.weekday()
    WEEKDAY_LABELS = ['月曜日', '火曜日', '水曜日', '木曜日', '金曜日', '土曜日', '日曜日']

    scheduled_customer_ids = _get_scheduled_customer_ids(target_date)
    orders_for_date = Order.objects.filter(
        delivery_date=target_date
    ).select_related('customer__delivery_bin').prefetch_related('items')
    order_map = {o.customer_id: o for o in orders_for_date}

    all_customer_ids = scheduled_customer_ids | set(order_map.keys())
    all_customers = Customer.objects.filter(
        pk__in=all_customer_ids, is_active=True
    ).select_related('delivery_bin')

    bins = DeliveryBin.objects.filter(is_active=True)
    assigned_bin_ids = {db.pk for db in bins}

    completed_set = set(
        DeliveryCompletion.objects.filter(delivery_date=target_date)
        .values_list('customer_id', flat=True)
    )

    bin_sections = []
    for db in bins:
        bin_customers = [c for c in all_customers if c.delivery_bin_id == db.pk]
        ordered = sum(1 for c in bin_customers if c.pk in order_map)
        total_qty = sum(
            order_map[c.pk].total_quantity for c in bin_customers if c.pk in order_map
        )
        bin_sections.append({
            'bin': db,
            'total_customers': len(bin_customers),
            'ordered_count': ordered,
            'total_quantity': total_qty,
            'completed_count': sum(1 for c in bin_customers if c.pk in completed_set),
        })

    unassigned_customers = [
        c for c in all_customers
        if c.delivery_bin_id is None or c.delivery_bin_id not in assigned_bin_ids
    ]
    unassigned_section = {
        'total_customers': len(unassigned_customers),
        'ordered_count': sum(1 for c in unassigned_customers if c.pk in order_map),
        'total_quantity': sum(
            order_map[c.pk].total_quantity for c in unassigned_customers if c.pk in order_map
        ),
    }

    context = {
        'target_date': target_date,
        'target_date_str': target_date.strftime('%Y-%m-%d'),
        'weekday_label': WEEKDAY_LABELS[weekday],
        'bin_sections': bin_sections,
        'unassigned_section': unassigned_section,
        'today': today,
    }
    return render(request, 'orders/delivery_list.html', context)


@login_required
def delivery_list_by_bin(request, pk):
    """特定便の配達リスト（配達員向け・モバイルファースト）"""
    delivery_bin = get_object_or_404(DeliveryBin, pk=pk, is_active=True)
    target_date, today = _parse_target_date(request)
    weekday = target_date.weekday()
    WEEKDAY_LABELS = ['月曜日', '火曜日', '水曜日', '木曜日', '金曜日', '土曜日', '日曜日']

    scheduled_customer_ids = _get_scheduled_customer_ids(target_date)
    orders_for_date = Order.objects.filter(
        delivery_date=target_date,
        customer__delivery_bin=delivery_bin
    ).select_related('customer', 'customer__payment_method').prefetch_related('items')
    order_map = {o.customer_id: o for o in orders_for_date}

    # この便に属する顧客（スケジュール済み + 当日受注済み）
    all_customer_ids = (
        scheduled_customer_ids | set(order_map.keys())
    )
    bin_customers = Customer.objects.filter(
        pk__in=all_customer_ids,
        is_active=True,
        delivery_bin=delivery_bin
    ).select_related('payment_method').order_by('company_name', 'name')

    completed_ids_bin = set(
        DeliveryCompletion.objects.filter(
            delivery_date=target_date,
            delivery_bin=delivery_bin
        ).values_list('customer_id', flat=True)
    )

    rows = []
    for customer in bin_customers:
        order = order_map.get(customer.pk)
        rows.append({
            'customer': customer,
            'order': order,
            'has_order': order is not None,
            'total_quantity': order.total_quantity if order else 0,
            'notes': order.notes if order else customer.notes,
            'is_completed': customer.pk in completed_ids_bin,
        })

    ordered_count = sum(1 for r in rows if r['has_order'])

    # 支払い方法別合計
    pm_map = {}
    for row in rows:
        if not row['has_order']:
            continue
        pm = row['customer'].payment_method
        key = pm.pk if pm else 0
        if key not in pm_map:
            pm_map[key] = {
                'name': pm.name if pm else 'その他',
                'total': 0,
                'sort_order': pm.sort_order if pm else 999,
            }
        pm_map[key]['total'] += row['order'].total
    payment_method_totals = sorted(pm_map.values(), key=lambda x: x['sort_order'])

    public_url = request.build_absolute_uri(
        f"/orders/public/delivery/{pk}/"
    )
    context = {
        'delivery_bin': delivery_bin,
        'target_date': target_date,
        'target_date_str': target_date.strftime('%Y-%m-%d'),
        'weekday_label': WEEKDAY_LABELS[weekday],
        'rows': rows,
        'ordered_count': ordered_count,
        'not_ordered_count': len(rows) - ordered_count,
        'payment_method_totals': payment_method_totals,
        'today': today,
        'public_url': public_url,
    }
    return render(request, 'orders/delivery_list_by_bin.html', context)


def delivery_list_by_bin_public(request, pk):
    """特定便の配達リスト（公開版・ログイン不要・QRコードアクセス対応）"""
    delivery_bin = get_object_or_404(DeliveryBin, pk=pk, is_active=True)
    target_date, today = _parse_target_date(request)
    weekday = target_date.weekday()
    WEEKDAY_LABELS = ['月曜日', '火曜日', '水曜日', '木曜日', '金曜日', '土曜日', '日曜日']

    scheduled_customer_ids = _get_scheduled_customer_ids(target_date)
    orders_for_date = Order.objects.filter(
        delivery_date=target_date,
        customer__delivery_bin=delivery_bin
    ).select_related('customer', 'customer__payment_method').prefetch_related(
        'items', 'extra_items'
    )
    order_map = {o.customer_id: o for o in orders_for_date}

    all_customer_ids = (
        scheduled_customer_ids | set(order_map.keys())
    )
    bin_customers = Customer.objects.filter(
        pk__in=all_customer_ids,
        is_active=True,
        delivery_bin=delivery_bin
    ).select_related('payment_method').order_by('company_name', 'name')

    completed_ids = set(
        DeliveryCompletion.objects.filter(
            delivery_date=target_date,
            delivery_bin=delivery_bin
        ).values_list('customer_id', flat=True)
    )

    rows = []
    for customer in bin_customers:
        order = order_map.get(customer.pk)
        rows.append({
            'customer': customer,
            'order': order,
            'has_order': order is not None,
            'total_quantity': order.total_quantity if order else 0,
            'notes': order.notes if order else customer.notes,
            'order_items': list(order.items.all()) if order else [],
            'extra_items': list(order.extra_items.all()) if order else [],
            'is_completed': customer.pk in completed_ids,
        })

    ordered_count = sum(1 for r in rows if r['has_order'])

    pm_map = {}
    for row in rows:
        if not row['has_order']:
            continue
        pm = row['customer'].payment_method
        key = pm.pk if pm else 0
        if key not in pm_map:
            pm_map[key] = {
                'name': pm.name if pm else 'その他',
                'total': 0,
                'sort_order': pm.sort_order if pm else 999,
            }
        pm_map[key]['total'] += row['order'].total
    payment_method_totals = sorted(pm_map.values(), key=lambda x: x['sort_order'])

    context = {
        'delivery_bin': delivery_bin,
        'target_date': target_date,
        'target_date_str': target_date.strftime('%Y-%m-%d'),
        'weekday_label': WEEKDAY_LABELS[weekday],
        'rows': rows,
        'ordered_count': ordered_count,
        'not_ordered_count': len(rows) - ordered_count,
        'payment_method_totals': payment_method_totals,
    }
    return render(request, 'orders/delivery_list_by_bin_public.html', context)


@csrf_exempt
def api_toggle_delivery_complete(request):
    """配達完了トグルAPI（公開・認証不要・QRコードアクセス対応）"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    try:
        data = json.loads(request.body)
        customer_id = int(data['customer_id'])
        delivery_date_str = data['delivery_date']
        bin_pk = int(data['bin_pk'])
        delivery_date = datetime.strptime(delivery_date_str, '%Y-%m-%d').date()
    except (KeyError, ValueError, TypeError):
        return JsonResponse({'error': 'Invalid parameters'}, status=400)

    customer = get_object_or_404(Customer, pk=customer_id)
    delivery_bin = get_object_or_404(DeliveryBin, pk=bin_pk)

    obj, created = DeliveryCompletion.objects.get_or_create(
        customer=customer,
        delivery_date=delivery_date,
        defaults={'delivery_bin': delivery_bin}
    )
    if not created:
        obj.delete()
        return JsonResponse({'completed': False})
    from django.utils import timezone as tz
    local_time = tz.localtime(obj.completed_at)
    return JsonResponse({'completed': True, 'completed_at': local_time.strftime('%H:%M')})
