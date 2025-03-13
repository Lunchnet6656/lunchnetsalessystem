# lunchnetsale/views.py
from django.shortcuts import render, redirect, get_object_or_404
from datetime import datetime, timedelta, date
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from .forms import UploadFileForm, UploadMenuForm, UploadItemQuantityForm, ProductForm, ItemQuantityForm, DailyReportForm, DailyReportEntryForm, TimeForm,ShiftRequestForm
import pandas as pd
from sales.models import SalesLocation, Product, ItemQuantity, DailyReport, DailyReportEntry, CustomUser, OthersItem,ShiftRequest,ShiftRequest,Holiday
import openpyxl
from django.db.models import Sum, Max, Count, Avg, Q
from django.utils.dateparse import parse_date
from django.utils.dateformat import format
from django.utils import timezone
from collections import defaultdict
from django.http import JsonResponse
from django.urls import reverse
import csv
import calendar
from calendar import monthrange
from dateutil.relativedelta import relativedelta
from django.utils.timezone import make_aware
import logging
from django.contrib.auth import get_user_model
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.forms import UserChangeForm
from django.views.decorators.http import require_POST
import json
from django import forms
from django.contrib.auth import update_session_auth_hash
from django.db.models import OuterRef, Subquery
import calendar as django_calendar  # Pythonの標準ライブラリからインポート
import qrcode
from io import BytesIO
import base64
from .forms import ShiftSubmissionForm
from datetime import date as dt_date
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.timezone import now
import urllib.parse

@login_required
def my_page(request):
    user = request.user  # 現在ログインしているユーザーを取得
    submitted_shifts = ShiftRequest.objects.filter(user=user).order_by('date')  # 提出済みシフトを取得
        # 会社の休日を取得
    holidays = Holiday.objects.values_list('date', flat=True)
    holidays_json = json.dumps([str(holiday) for holiday in holidays])  # JSON形式に変換

    # 変更不可日を判定（5日と20日）
    today = dt_date.today()
    can_edit_shifts = today.day <= 5 or today.day <= 20

    # 勤怠用QRコードの生成
    qr = qrcode.make(request.user.username)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    qr_code = base64.b64encode(buffer.getvalue()).decode()

    return render(request, 'mypage.html', {
        'shifts': submitted_shifts,
        'can_edit_shifts': can_edit_shifts,  # 変更可能かどうかをテンプレートに渡す
        'qr_code': qr_code,
        'holidays_json':holidays_json,
    })

@login_required
def submit_shift(request):
    user = request.user
    today = dt_date.today()

    # 会社の休日を取得
    holidays = Holiday.objects.values_list('date', flat=True)
    holidays_json = json.dumps([str(holiday) for holiday in holidays])  # JSON形式に変換

    if request.method == "POST":
        holidays_json = json.dumps([str(holiday) for holiday in holidays])  # JSON形式に変換
        form = ShiftSubmissionForm(request.POST)
        form.initial['user'] = user  # バリデーション用にユーザー情報をセット

        if form.is_valid():
            shift_requests = []

            for date, is_off in form.cleaned_data.items():
                if date.startswith("status_"):  # status_YYYY-MM-DD のデータを処理
                    shift_date = dt_date.fromisoformat(date.replace("status_", ""))

                    shift_requests.append(
                        ShiftRequest(
                            user=user,
                            first_name=user.first_name,
                            last_name=user.last_name,
                            date=shift_date,
                            is_off=is_off,
                            comment=request.POST.get('message', ''),
                        )
                    )

                    # 既存データのチェック
                    if ShiftRequest.objects.filter(user=user, date=shift_date).exists():
                        messages.error(request, f"{shift_date} からのシフトは既に送信されています。\nマイページのシフト編集から変更してください。")
                        return redirect('submit_shift')

            if shift_requests: 
                ShiftRequest.objects.bulk_create(shift_requests)  # まとめて保存
                messages.success(request, "シフト希望が送信されました")
                return redirect("mypage")
            else:
                messages.error(request, "エラーがあります。")
    
    else:
        form = ShiftSubmissionForm()

    return render(request, 'submit_shift.html', {
        'form': form,
        'today': today,
        'holidays_json': holidays_json,  # JavaScript に休日データを渡す
    })

@login_required
def edit_shift(request):
    today = dt_date.today()
    # 編集可能かどうかを判定
    can_edit_shifts = today.day <= 5 or (today.day > 5 and today.day <= 20)

    # 会社の休日を取得
    holidays = Holiday.objects.values_list('date', flat=True)
    holidays_json = json.dumps([str(holiday) for holiday in holidays])  # JSON形式に変換

    if request.method == 'POST':
        if not can_edit_shifts:
            messages.error(request, "シフトの締切日を過ぎているため、編集できません。")
            return redirect('mypage')
        
        holidays_json = json.dumps([str(holiday) for holiday in holidays])  # JSON形式に変換

        # ユーザーの全シフトを取得
        shifts = ShiftRequest.objects.filter(user=request.user)
        # コメントをPOSTデータから取得
        comment = request.POST.get('comment')  # 同じコメントを全シフトに適用
        for shift in shifts:
            # POSTデータから対応するシフトに関連する情報を取得
            shift_status = request.POST.get(f'shift_{shift.id}')  # チェックボックスの値を取得

            # 'shift_status' が 'on' の場合は休み希望、そうでなければ出勤に設定
            if shift_status == 'on':
                shift.is_off = True
            else:
                shift.is_off = False

            # コメントを更新
            shift.comment = comment  # すべてのシフトに同じコメントを設定

            # 変更を保存
            shift.save()
        messages.success(request, "シフト希望が編集されました")
        return redirect('mypage')  # 編集後、マイページにリダイレクト

    # GETリクエストのときは、ユーザーのシフトを表示
    shifts = ShiftRequest.objects.filter(user=request.user)  # ユーザーのシフトを取得
    return render(request, 'edit_shift.html', {
        'shifts': shifts,
        'holidays_json': holidays_json,  # JavaScript に休日データを渡す
    })

@staff_member_required
def shift_settings(request):
    today = now().date()
    
    # URLのパラメータから年月を取得（デフォルトは今月）
    year = request.GET.get("year", today.year)
    month = request.GET.get("month", today.month)
    
    try:
        year = int(year)
        month = int(month)
    except ValueError:
        year = today.year
        month = today.month

    # 表示する月の開始・終了日
    start_date = datetime(year, month, 1).date()
    if month == 12:
        end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)

    # 既存の休日を取得
    existing_holidays = {holiday.date for holiday in Holiday.objects.filter(date__range=[start_date, end_date])}

    # カレンダー用データを作成
    calendar_data = []
    current_date = start_date
    while current_date <= end_date:
        calendar_data.append({
            'date': current_date,
            'is_weekend': current_date.weekday() in [5, 6],  # 土日
            'is_holiday': current_date in existing_holidays  # 既存の休日
        })
        current_date += timedelta(days=1)

    # 前後の月
    prev_month = (year, month - 1) if month > 1 else (year - 1, 12)
    next_month = (year, month + 1) if month < 12 else (year + 1, 1)

    if request.method == "POST":
        # 選択された休日データを取得
        selected_dates = request.POST.getlist("holiday_dates")
        new_holidays = set(datetime.strptime(x, "%Y-%m-%d").date() for x in selected_dates)

        # 現在の休日を更新
        Holiday.objects.filter(date__range=[start_date, end_date]).exclude(date__in=new_holidays).delete()
        for date in new_holidays:
            Holiday.objects.get_or_create(date=date)

        messages.success(request, "カレンダーの設定を変更しました。")
        return redirect(request.path + f"?year={year}&month={month}")

    return render(request, "shift_settings.html", {
        "calendar_data": calendar_data,
        "current_month": start_date,
        "prev_month": prev_month,
        "next_month": next_month,
    })


@staff_member_required
def shift_request_list(request):
    # ユーザーごとに最新のShiftRequestを取得
    latest_shifts = ShiftRequest.objects.values('user').annotate(latest_submission=Max('submitted_at'))
    # 最新のShiftRequestの詳細情報を取得
    shifts = ShiftRequest.objects.filter(submitted_at__in=[entry['latest_submission'] for entry in latest_shifts])



    return render(request, 'shift_request_list.html', {'shifts': shifts})

@staff_member_required
def shift_request_detail(request, pk):
    shift_request = get_object_or_404(ShiftRequest, pk=pk)
    # 会社の休日を取得
    holidays = Holiday.objects.values_list('date', flat=True)
    holidays_json = json.dumps([str(holiday) for holiday in holidays])  # JSON形式に変換

    # シフトリクエストの詳細情報を取得
    user = shift_request.user
    shifts = ShiftRequest.objects.filter(user=user).order_by('date')

    return render(request, 'shift_request_detail.html', {
        'shift_request': shift_request,
        'shifts': shifts,
        'holidays_json': holidays_json,  # JavaScript に休日データを渡す
    })



@staff_member_required
def export_shift_request(request):
    shifts = ShiftRequest.objects.all()  # すべてのShiftRequestを取得
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="shift_requests.csv"'
    writer = csv.writer(response)
    
    # CSVヘッダー
    writer.writerow(['名前', '日付', '出勤状態', '伝言'])

    # CSVデータ
    for shift in shifts:
        writer.writerow([f"{shift.last_name} {shift.first_name}", shift.date, '休み' if shift.is_off else '出勤', shift.comment])

    return response


def login_view(request):
    logger = logging.getLogger(__name__)
    User = get_user_model()

    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        logger.info(f'Trying to authenticate user: {username}')

        # ユーザーが存在するか確認
        try:
            user = User.objects.get(username=username)
            logger.info(f'User found: {user.username}')
        except User.DoesNotExist:
            logger.error(f'User not found: {username}')
            messages.error(request, 'ユーザーが存在しません')
            return render(request, 'login.html')

        # 認証を試行
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            logger.error(f'Authentication failed for user: {username}')
            messages.error(request, 'ユーザー名またはパスワードが正しくありません')

    return render(request, 'login.html')

def locked_out_view(request):
    return render(request, 'axes/locked_out.html')

def logout_view(request):
    logout(request)
    messages.info(request, 'ログアウトしました。')  # メッセージを設定
    return redirect('login')  # リダイレクト先のURL名を指定

@login_required
def dashboard_view(request):

    # セッションの有効期限を確認
    if request.session.get_expiry_age() <= 0:  # セッション期限が切れている場合
        messages.warning(request, 'セッションの有効期限が切れました。再度ログインしてください。')
        request.session.flush()  # セッションデータをクリア
        return redirect('login')  # 'login' はログインページのURL名前付きルート

    today = datetime.today()#date(2025, 1, 15)

    # 本日の売上、販売数、残数を集計
    daily_reports = DailyReport.objects.filter(date=today).order_by('location_no')

    total_quantity = sum(report.total_quantity for report in daily_reports)
    total_revenue = sum(report.total_revenue for report in daily_reports)
    total_sales_quantity = sum(report.total_sales_quantity for report in daily_reports)
    total_remaining = sum(report.total_remaining for report in daily_reports)
    # Waste_rateの計算
    if total_quantity > 0:  # ゼロ除算を防ぐためのチェック
        waste_rate = (total_remaining / total_quantity) * 100  # パーセンテージに変換
    else:
        waste_rate = 0.0

    # 小数点以下2位までフォーマット
    formatted_waste_rate = f"{waste_rate:.1f}%"


    # メニュー別の集計
    menu_summary = DailyReportEntry.objects.filter(report__date=today).values('product','product_no').annotate(
        total_quantity=Sum('quantity'),
        total_sales_quantity=Sum('sales_quantity'),
        total_remaining=Sum('remaining_number')
    ).order_by('product_no')

    context = {
        'today': today,
        'total_quantity': total_quantity,
        'total_revenue': total_revenue,
        'total_sales_quantity': total_sales_quantity,
        'total_remaining': total_remaining,
        'formatted_waste_rate': formatted_waste_rate,
        'menu_summary': menu_summary,
        'report':daily_reports,
    }


    return render(request, 'dashboard.html', context)

@login_required
def user_list_view(request):
    users = User.objects.all().order_by('date_joined')  # ユーザー情報を古い順に取得
    return render(request, 'user_list.html', {'users': users})


# カスタムフォームを作成
class CustomUserChangeForm(UserChangeForm):
    current_password = forms.CharField(
        required=False,
        widget=forms.PasswordInput(attrs={'placeholder': '現在のパスワード'}),
        label='現在のパスワード'
    )
    new_password = forms.CharField(
        required=False,
        widget=forms.PasswordInput(attrs={'placeholder': '新しいパスワード'}),
        label='新しいパスワード'
    )
    confirm_password = forms.CharField(
        required=False,
        widget=forms.PasswordInput(attrs={'placeholder': '新しいパスワード（確認）'}),
        label='新しいパスワード（確認）'
    )

    class Meta:
        model = User
        fields = ['username', 'first_name', 'last_name', 'is_staff']

    def clean(self):
        cleaned_data = super().clean()
        new_password = cleaned_data.get('new_password')
        confirm_password = cleaned_data.get('confirm_password')

        if new_password and new_password != confirm_password:
            raise forms.ValidationError('新しいパスワードと確認用パスワードが一致しません。')

        return cleaned_data
    
@login_required
def user_edit_view(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=user)
        current_password = request.POST.get('current_password', '').strip()
        new_password = request.POST.get('new_password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()

        if form.is_valid():
            # パスワード変更処理が必要か確認
            if current_password or new_password or confirm_password:
                # 現在のパスワードが入力されている場合のみチェック
                if not current_password:
                    form.add_error('current_password', '現在のパスワードを入力してください。')
                elif not user.check_password(current_password):
                    form.add_error('current_password', '現在のパスワードが正しくありません。')
                elif new_password or confirm_password:  # 新しいパスワードのチェック
                    if new_password != confirm_password:
                        form.add_error('new_password', '新しいパスワードが一致しません。')
                    elif new_password:
                        user.set_password(new_password)

            # フォームがエラーを含む場合、保存処理を中断
            if form.errors:
                return render(request, 'user_edit.html', {'form': form})

            # 他のフィールドを保存
            user.save()

            # パスワードが変更された場合、セッションを更新
            if new_password:
                update_session_auth_hash(request, user)

            messages.success(request, f'ユーザー {user.username} を更新しました。')
            return redirect('user_list')
    else:
        form = CustomUserChangeForm(instance=user)

    return render(request, 'user_edit.html', {'form': form})

@login_required
def user_delete_view(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.delete()
        messages.success(request, f'ユーザー {user.username} を削除しました。')
        return JsonResponse({'message': 'ユーザーを削除しました。'})
    return render(request, 'user_confirm_delete.html', {'user': user})


@login_required
def upload_view(request):
    if request.method == 'POST':
        location_form = UploadFileForm(request.POST, request.FILES)
        menu_form = UploadMenuForm(request.POST, request.FILES)
        item_quantity_form = UploadItemQuantityForm(request.POST, request.FILES)

        if 'location_file' in request.FILES and location_form.is_valid():
            file = request.FILES['location_file']
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                no, name, type, price_type, service_name, service_price, direct_return = row
                # service_priceがNoneまたは空であればデフォルト値を設定
                if service_price is None or service_price == '':
                    service_price = 0  # デフォルト値を設定
                SalesLocation.objects.create(no=no, name=name, type=type, price_type=price_type, service_name=service_name, service_price=service_price, direct_return=direct_return)
            messages.success(request, '販売場所データがアップロードされました。')

        elif 'menu_file' in request.FILES and menu_form.is_valid():
            file = request.FILES['menu_file']
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                no, week, name, price_A, price_B, price_C, container_type = row
                week = datetime.strptime(str(week), "%Y%m%d").date()

                # weekの重複チェック
                product_exists = Product.objects.filter(week=week, no=no).exists()
                if product_exists:
                    # 既存データがあれば、更新
                    product = Product.objects.get(week=week, no=no)
                    product.name = name
                    product.price_A = price_A
                    product.price_B = price_B
                    product.price_C = price_C
                    product.container_type = container_type
                    product.save()
                else:
                    # 新規作成
                    Product.objects.create(no=no, week=week, name=name, price_A=price_A, price_B=price_B, price_C=price_C, container_type=container_type)

            messages.success(request, 'メニューデータがアップロードされました。')

        elif 'item_quantity_file' in request.FILES and item_quantity_form.is_valid():
            file = request.FILES['item_quantity_file']
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            target_week = sheet.cell(row=2, column=1).value
            target_week = datetime.strptime(str(target_week), "%Y%m%d").date()
            date = datetime.strptime(sheet.title, "%Y%m%d").date()

            for row in sheet.iter_rows(min_row=2, max_row=12, values_only=True):
                product_no = row[1]
                try:
                    product = Product.objects.get(no=product_no, week=target_week)
                except Product.DoesNotExist:
                    continue

                for col_idx, quantity in enumerate(row[2:], start=2):
                    sales_location_name = sheet.cell(row=1, column=col_idx + 1).value
                    try:
                        sales_location = SalesLocation.objects.get(name=sales_location_name)
                    except SalesLocation.DoesNotExist:
                        continue

                    # target_weekとtarget_dateが重複しているか確認
                    item_quantity_exists = ItemQuantity.objects.filter(
                        target_week=target_week, target_date=date, product=product, sales_location=sales_location
                    ).exists()

                    if item_quantity_exists:
                        # 既存データがあれば更新
                        item_quantity = ItemQuantity.objects.get(
                            target_week=target_week, target_date=date, product=product, sales_location=sales_location
                        )
                        item_quantity.quantity = quantity
                        item_quantity.save()
                    else:
                        # 新規作成
                        ItemQuantity.objects.create(
                            target_date=date,
                            target_week=target_week,
                            product=product,
                            sales_location=sales_location,
                            quantity=quantity
                        )
            messages.success(request, '持参数データがアップロードされました。')

        else:
            messages.error(request, 'フォームにエラーがあります。')
    else:
        location_form = UploadFileForm()
        menu_form = UploadMenuForm()
        item_quantity_form = UploadItemQuantityForm()

    return render(request, 'upload.html', {
        'location_form': location_form,
        'menu_form': menu_form,
        'item_quantity_form': item_quantity_form
    })

@login_required
def register_user(request):
    if request.method == 'POST':
        username = request.POST['username']
        password1 = request.POST['password1']
        password2 = request.POST['password2']
        user_role = request.POST['user_role']
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']

        if password1 != password2:
            return render(request, 'register.html', {'error': 'パスワードが一致しません'})

        try:
            user = User.objects.create_user(username=username, password=password1, first_name=first_name, last_name=last_name)
            # 権限の設定
            if user_role == 'admin':
                user.is_staff = True
            else:
                user.is_staff = False
            user.save()

            success_message = 'ユーザーの登録が完了しました。'
            return render(request, 'register.html', {'success_message': success_message})

        except Exception as e:
            return render(request, 'register.html', {'error': str(e)})

    return render(request, 'register.html')


# ヘルパー関数の定義
def parse_value(value_str):
    """
    文字列形式の数値または割引額を数値に変換します。
    - 数値の場合、カンマを取り除いて整数に変換します。
    - 割引額の場合、'▲'、'±'、'+' の記号を処理して整数に変換します。
    """
    if not value_str:
        return 0

    # 割引額の前に付く記号を処理
    if value_str.startswith('▲'):
        # '▲'の場合、マイナス記号を追加して数値を変換
        value_str = '-' + value_str[1:]
    elif value_str.startswith('±') or value_str.startswith('+'):
        # '±'や'+'の場合、記号を取り除いて数値に変換
        value_str = value_str[1:]

    # カンマを取り除く
    value_str = value_str.replace(',', '')

    # 数値に変換
    try:
        return int(value_str)
    except ValueError:
        # 数値に変換できない場合は0を返す
        return 0


@login_required
def daily_report_view(request):

    selected_person = ""
    # ログインユーザーのfirst_nameを取得
    selected_person = request.user.first_name if request.user.is_authenticated else ""
    print(f"Logged-in user's first name: {selected_person}")  # デバッグ用

    # 日付と販売場所の取得
    dates = ItemQuantity.objects.values_list('target_date', flat=True).distinct().order_by('-target_date')[:3]
    locations = SalesLocation.objects.all()
    others_item = OthersItem.objects.all() 
    weather_options = ["快晴", "晴れ", "くもり", "雨", "大雨", "雪"]
    temp_options = ["猛暑", "暑い", "ちょうどいい", "涼しい", "寒い",'極寒']

    # 初期化
    selected_date = None
    selected_location = None
    selected_weather = []
    selected_temp = []
    item_data = []
    service = None



    if request.method == 'POST':
        action = request.POST.get('action') # ボタンを判断する。
        print(f"送信されたPOSTデータ: {request.POST}")  # ここで全POSTデータを確認
        selected_date = request.POST.get('date')
        selected_location = request.POST.get('location')
        selected_person = request.POST.get('担当者',  f"{request.user.last_name} {request.user.first_name}")
        print(f"送信される担当者名: {selected_person}")  # ここで値を確認
        print(f"Logged-in user's first name: {selected_person}")
        selected_weather = request.POST.getlist('weather')
        selected_temp = request.POST.getlist('temp')

        # 売上関連のデータ
        total_quantity = request.POST.get('total_quantity',0)
        total_sales_quantity = request.POST.get('total_sales_quantity',0)
        total_remaining = request.POST.get('total_remaining',0)
        total_sales = parse_value(request.POST.get('total_sales', 0))
        total_revenue = parse_value(request.POST.get('total_revenue', 0))
        sales_price_quantity_1 = request.POST.get('sales_price_quantity_1',0)
        sales_price_quantity_2 = request.POST.get('sales_price_quantity_2',0)
        sales_price_quantity_3 = request.POST.get('sales_price_quantity_3',0)

        # その他の売上関連のデータ
        others_sales_1 = request.POST.get('selected_item_1', 0)
        others_price1 = request.POST.get('others_price1', 0)
        others_sales_quantity1 = request.POST.get('others_sales_quantity1', 0)
        others_sales_2 = request.POST.get('selected_item_2', 0)
        others_price2 = request.POST.get('others_price2', 0)
        others_sales_quantity2 = request.POST.get('others_sales_quantity2', 0)
        total_others_sales = parse_value(request.POST.get('total_others_sales', 0))

        # 割引関連のデータ
        no_rice_quantity = request.POST.get('no_rice_quantity',0)
        extra_rice_quantity = request.POST.get('extra_rice_quantity',0)
        coupon_type_600 = request.POST.get('coupon_type_600',0)
        coupon_type_700 = request.POST.get('coupon_type_700',0)
        discount_50 = request.POST.get('discount_50',0)
        discount_100 = request.POST.get('discount_100',0)
        # 割引関連のデータ(サービス)
        service_name = request.POST.get('service_name', 0)
        service_price = request.POST.get('service_price', 0)
        service_type_600 = request.POST.get('service_type_600', 0)
        service_type_700 = request.POST.get('service_type_700', 0)
        service_type_100 = request.POST.get('service_type_100', 0)
        total_discount = parse_value(request.POST.get('total_discount', '0'))

        # 売上金関連のデータ
        paypay = request.POST.get('paypay', 0)
        digital_payment = request.POST.get('digital_payment', 0)
        cash = request.POST.get('cash', 0)
        sales_difference = parse_value(request.POST.get('sales_difference', '0'))

        # 時間データの取得
        departure_time = request.POST.get('departure_time')
        arrival_time = request.POST.get('arrival_time')
        opening_time = request.POST.get('opening_time')
        sold_out_time = request.POST.get('sold_out_time')
        closing_time = request.POST.get('closing_time')

        # 経費関連のデータ
        gasolin = request.POST.get('gasolin', 0)
        highway = request.POST.get('highway', 0)
        parking = request.POST.get('parking', 0)
        part = request.POST.get('part', 0)
        others = request.POST.get('others', 0)

        # 情報関連のデータ
        comment = request.POST.get('comment', 0)
        food_count_setting = request.POST.get('food_count_setting', 0)

        # 日付と販売場所に基づいてデータを取得
        item_quantities = ItemQuantity.objects.filter(
            target_date=selected_date,
            sales_location__name=selected_location
        ).select_related('product', 'sales_location')


        service = SalesLocation.objects.get(name=selected_location)
        print(service.service_name)
        print(service.service_price)
        
        # 商品の情報を取得
        products = Product.objects.filter(
            no__in=item_quantities.values_list('product__no', flat=True),
            itemquantity__target_date=selected_date  # target_dateを条件に追加
        )
        print(f"デバッグ : {item_quantities} ")
        
        # 商品情報を辞書にまとめる
        product_data = {
            product.no: product for product in products
        }

        # 各商品の持参数を計算
        item_data = []
        unique_prices = set()  # ユニークな価格を格納するセット
        for item in item_quantities:
            product = product_data.get(item.product.no)
            if product:
                # SalesLocationのprice_typeに応じて価格を設定
                if item.sales_location.price_type == 'A':
                    price = product.price_A
                elif item.sales_location.price_type == 'B':
                    price = product.price_B
                elif item.sales_location.price_type == 'C':
                    price = product.price_C
                else:
                    price = 0  # デフォルト値として0などを設定

                unique_prices.add(price)  # ユニークな価格を追加

                item_data.append({
                    'menu_no': product.no,
                    'menu_name': product.name,
                    'price': price,
                    'quantity': item.quantity,
                    'remaining': item.quantity,
                    'total_sales': 0
                })
        # unique_pricesをリストに変換し、降順にソート
        unique_prices = sorted(list(unique_prices), reverse=True)

        # menu_noでソート
        item_data = sorted(item_data, key=lambda x: x['menu_no'])
        
        # 日付と販売場所に基づいて既存のDailyReportを取得または作成
        if action == 'send':
            print(f"送信される担当者名: {selected_person}")  # ここで値を確認
            
            # 既存のレポートを確認
            existing_report = DailyReport.objects.filter(
                date=parse_date(selected_date),
                location=selected_location
            ).first()
            
            # confirmedがTrueの場合は処理を中止
            if existing_report and existing_report.confirmed:
                messages.error(request, f'この日付（{selected_date}）の{selected_location}の日計表は既に集計済みのため、送信できません。', extra_tags='alert alert-danger')
                return redirect('daily_report')
                
            # 既存の処理を続行
            report, created = DailyReport.objects.update_or_create(
                date=parse_date(selected_date),
                location=selected_location,
                defaults={
                    'location_no':service.no,
                    'person_in_charge': selected_person,
                    'weather': ','.join(selected_weather),
                    'temp': ','.join(selected_temp),
                    'total_quantity': total_quantity,
                    'sales_price_quantity_1': sales_price_quantity_1,
                    'sales_price_quantity_2': sales_price_quantity_2,
                    'sales_price_quantity_3': sales_price_quantity_3, 
                    'total_sales_quantity': total_sales_quantity,
                    'total_remaining': total_remaining,
                    'total_revenue': total_revenue,
                    'others_sales_1': others_sales_1,
                    'others_price1': others_price1,
                    'others_sales_quantity1': others_sales_quantity1,
                    'others_sales_2': others_sales_2,
                    'others_price2': others_price2,
                    'others_sales_quantity2': others_sales_quantity2,
                    'total_others_sales': total_others_sales,
                    'no_rice_quantity': no_rice_quantity,
                    'extra_rice_quantity': extra_rice_quantity,
                    'coupon_type_600': coupon_type_600,
                    'coupon_type_700': coupon_type_700,
                    'discount_50': discount_50,
                    'discount_100': discount_100,
                    'service_name': service_name,
                    'service_price': service_price,
                    'service_type_600': service_type_600,
                    'service_type_700': service_type_700,
                    'service_type_100': service_type_100,
                    'total_discount': total_discount,
                    'paypay': paypay,
                    'digital_payment': digital_payment,
                    'cash': cash,
                    'sales_difference': sales_difference,
                    'departure_time': departure_time if departure_time else None,
                    'arrival_time': arrival_time if arrival_time else None,
                    'opening_time': opening_time if opening_time else None,
                    'sold_out_time': sold_out_time if sold_out_time else None,
                    'closing_time': closing_time if closing_time else None,
                    'gasolin': gasolin,
                    'highway': highway,
                    'parking': parking,
                    'part': part,
                    'others': others,
                    'comments': comment,
                    'food_count_setting': food_count_setting
                }
            )
            # DailyReportEntryの作成または更新
            for item in item_data:
                DailyReportEntry.objects.update_or_create(
                    report=report,
                    product=item.get('menu_name', ''),
                    product_no=item.get('menu_no', ''),
                    defaults={
                        'quantity': request.POST.get(f'quantity_{item["menu_no"]}', 0),
                        'sales_quantity': request.POST.get(f'sales_quantity_{item["menu_no"]}', 0),
                        'remaining_number': request.POST.get(f'remaining_{item["menu_no"]}', 0),
                        'total_sales': parse_value(request.POST.get(f'total_sales_{item["menu_no"]}', '0')),
                        'sold_out': request.POST.get(f'sold_out_{item["menu_no"]}', 'off') == 'on',
                        'popular': request.POST.get(f'popular_{item["menu_no"]}', 'off') == 'on',
                        'unpopular': request.POST.get(f'unpopular_{item["menu_no"]}', 'off') == 'on'
                    }
                )
        else:
            context = {
                'dates': dates,
                'locations': locations,
                'item_data': item_data,
                'selected_date': selected_date,
                'selected_location': selected_location,
                'selected_person': selected_person,
                'selected_weather': selected_weather,
                'weather_options': weather_options,
                'temp_options': temp_options,
                'selected_temp': selected_temp,
                'othersitem': others_item,
                'service': service,
                'comment': comment,
                'food_count_setting': food_count_setting,
                'unique_prices': unique_prices
            }
            # 保存処理やバリデーションなどの処理をここに記述
            return render(request, 'daily_report.html', context)

        return redirect('submission_complete') # 完了ページにリダイレクト

    return render(request, 'daily_report.html', {
        'dates': dates,
        'locations': locations,
        'weather_options': weather_options,
        'temp_options': temp_options,
        'item_data': item_data,
        'service': service,
        'selected_person': selected_person  # first_nameを渡す
    })

@login_required
def submission_complete_view(request):
    return render(request, 'submission_complete.html')


# DailyReportリストを表示するビュー
@login_required
def daily_report_list(request):
    search_date_start = request.GET.get('search_date_start', '')
    search_date_end = request.GET.get('search_date_end', '')

    if search_date_start and search_date_end:
        # 範囲指定でフィルタリング
        daily_reports = DailyReport.objects.filter(date__range=[search_date_start, search_date_end])
    else:
        daily_reports = DailyReport.objects.all()

    # 日付ごとにレポートを集約
    aggregated_reports = (
        daily_reports
        .values('date')  # 日付ごとにグループ化
        .annotate(
            count=Count('id'),  # 各グループの件数
            latest_update=Max('updated_at')  # 最新の更新日時
        )
        .order_by('-date')
    )

    # 送信されたデータを取得
    submitted_locations = daily_reports.values_list('location', flat=True).distinct()
    all_locations = SalesLocation.objects.all()

    # 送信されていない場所を特定
    not_submitted_locations = [location.name for location in all_locations if location.name not in submitted_locations]


    context = {
        'aggregated_reports': aggregated_reports,
        'search_date_start': search_date_start,
        'search_date_end': search_date_end,
        'not_submitted_locations': not_submitted_locations,
        'submitted_count': len(submitted_locations),
        'total_locations': all_locations.count(),
    }
    return render(request, 'daily_report_list.html', context)

# Location別の詳細ビュー
@login_required
def daily_report_detail(request, date):
    # 文字列の日付をdatetimeオブジェクトに変換
    date = datetime.strptime(date, '%Y-%m-%d').date()  # 例: '2024-10-05'の形式

    reports_by_location = DailyReport.objects.filter(date=date).order_by('location_no')

    # 送信されたデータを取得
    submitted_locations = reports_by_location.values_list('location', flat=True).distinct()
    all_locations = SalesLocation.objects.all()

    # 送信されていない場所を特定
    not_submitted_locations = [location.name for location in all_locations if location.name not in submitted_locations]

    context = {
        'reports_by_location': reports_by_location,
        'date': date,
        'not_submitted_locations': not_submitted_locations,
        'submitted_count': len(submitted_locations),
        'total_locations': all_locations.count(),
    }
    return render(request, 'daily_report_detail.html', context)

@login_required
def daily_report_detail_rol(request):

    # 現在ログインしているユーザー名を取得
    current_user = f"{request.user.last_name} {request.user.first_name}"  # スペースで区切る


    # ログインユーザーのデータに絞り込む
    reports_by_location = DailyReport.objects.filter(person_in_charge=current_user).order_by('-date')


    context = {
        'reports_by_location': reports_by_location,
    }
    return render(request, 'daily_report_detail_rol.html', context)
# 編集ビュー
@login_required
def daily_report_edit(request, pk):
    report = get_object_or_404(DailyReport, pk=pk)
    entries = report.entries.all()  # 関連するエントリを取得
    locations = SalesLocation.objects.all()

    # locationの中から、report.locationと一致するものを探す
    for location in locations:
        if location.name == report.location:
            report.service_name = location.service_name
            report.service_price = location.service_price
            break  # 一致するlocationが見つかったらループを終了

    if request.method == "POST":
        # TimeFormのインスタンスを作成
        time_form = TimeForm(request.POST)

        post_data = request.POST.copy()
        if not post_data.get('service_name'):
            post_data['service_name'] = 'なし'
        if not post_data.get('service_price'):
            post_data['service_price'] = '0'

        form = DailyReportForm(post_data, instance=report)
        
        if form.is_valid() and time_form.is_valid():
            # 時間データを明示的に取得
            report.departure_time = request.POST.get('departure_time')
            report.arrival_time = request.POST.get('arrival_time')
            report.opening_time = request.POST.get('opening_time')
            report.sold_out_time = request.POST.get('sold_out_time')
            report.closing_time = request.POST.get('closing_time')
            
            # 他のフィールドの保存
            report = form.save(commit=False)
            report.save()

            # 日計表明細の処理
            for index in range(1, len(entries) + 1):
                product_name = request.POST.get(f'product_{index}')
                quantity = request.POST.get(f'quantity_{index}')
                sales_quantity = request.POST.get(f'sales_quantity_{index}')
                remaining_number = request.POST.get(f'remaining_number_{index}')
                total_sales = request.POST.get(f'total_sales_{index}')
                sold_out = request.POST.get(f'sold_out_{index}') is not None
                popular = request.POST.get(f'popular_{index}') is not None
                unpopular = request.POST.get(f'unpopular_{index}') is not None

                # 既存のエントリを更新
                if index <= len(entries):
                    entry = entries[index - 1]
                    entry.product = product_name
                    entry.quantity = int(quantity)  # 整数型に変換
                    entry.sales_quantity = int(sales_quantity)  # 整数型に変換
                    entry.remaining_number = int(remaining_number)  # 整数型に変換
                    entry.total_sales = int(total_sales)  # 整数型に変換
                    entry.sold_out = sold_out
                    entry.popular = popular
                    entry.unpopular = unpopular
                    entry.save()
                else:
                    # 新しいエントリを作成
                    DailyReportEntry.objects.create(
                        report=report,
                        product=product_name,
                        quantity=int(quantity),  # 整数型に変換
                        sales_quantity=int(sales_quantity),  # 整数型に変換
                        remaining_number=int(remaining_number),  # 整数型に変換
                        total_sales=int(total_sales),  # 整数型に変換
                        sold_out=sold_out,
                        popular=popular,
                        unpopular=unpopular,
                    )

            messages.success(request, "更新されました")  # メッセージを追加
            return redirect('daily_report_detail', date=report.date)  # 保存後にリダイレクト
        else:
            print(form.errors)  # エラーを表示
            print(time_form.errors)  # TimeFormのエラーも表示

    else:
        form = DailyReportForm(instance=report)
        time_form = TimeForm(initial={
            'departure_time': report.departure_time,
            'arrival_time': report.arrival_time,
            'opening_time': report.opening_time,
            'sold_out_time': report.sold_out_time,
            'closing_time': report.closing_time,
        })

    return render(request, 'daily_report_edit.html', {
        'form': form,
        'time_form': time_form,
        'report': report,
        'entries': entries
    })

# 編集ビュー
@login_required
def daily_report_edit_rol(request, pk):
    print("=== daily_report_edit_rol 開始 ===")
    report = get_object_or_404(DailyReport, pk=pk)
    entries = report.entries.all()  # 関連するエントリを取得
    locations = SalesLocation.objects.all()

    # locationの中から、report.locationと一致するものを探す
    for location in locations:
        if location.name == report.location:
            report.service_name = location.service_name
            report.service_price = location.service_price
            break  # 一致するlocationが見つかったらループを終了

    if request.method == "POST":
        print("=== POST処理開始 ===")
        time_form = TimeForm(request.POST)

        post_data = request.POST.copy()
        if not post_data.get('service_name'):
            post_data['service_name'] = 'なし'
        if not post_data.get('service_price'):
            post_data['service_price'] = '0'

        form = DailyReportForm(post_data, instance=report)
        
        print("POSTデータ:", request.POST)
        
        if form.is_valid() and time_form.is_valid():
            print("フォームのバリデーション成功")
            try:
                # 時間データを明示的に取得
                report.departure_time = request.POST.get('departure_time')
                report.arrival_time = request.POST.get('arrival_time')
                report.opening_time = request.POST.get('opening_time')
                report.sold_out_time = request.POST.get('sold_out_time')
                report.closing_time = request.POST.get('closing_time')
                
                print("時間データ設定完了")
                
                # フォームから直接保存
                report = form.save()
                print("レポート保存完了")

                # 日計表明細の処理
                print("=== エントリー処理開始 ===")
                for index in range(1, len(entries) + 1):
                    print(f"エントリー {index} 処理中")
                    product_name = request.POST.get(f'product_{index}')
                    quantity = request.POST.get(f'quantity_{index}')
                    sales_quantity = request.POST.get(f'sales_quantity_{index}')
                    remaining_number = request.POST.get(f'remaining_number_{index}')
                    total_sales = request.POST.get(f'total_sales_{index}')
                    sold_out = request.POST.get(f'sold_out_{index}') is not None
                    popular = request.POST.get(f'popular_{index}') is not None
                    unpopular = request.POST.get(f'unpopular_{index}') is not None

                    print(f"エントリーデータ: product={product_name}, quantity={quantity}, sales={sales_quantity}")

                    if index <= len(entries):
                        try:
                            entry = entries[index - 1]
                            entry.product = product_name
                            entry.quantity = int(quantity) if quantity else 0
                            entry.sales_quantity = int(sales_quantity) if sales_quantity else 0
                            entry.remaining_number = int(remaining_number) if remaining_number else 0
                            entry.total_sales = int(total_sales) if total_sales else 0
                            entry.sold_out = sold_out
                            entry.popular = popular
                            entry.unpopular = unpopular
                            entry.save()
                            print(f"エントリー {index} 保存成功")
                        except Exception as e:
                            print(f"エントリー {index} 保存エラー: {str(e)}")
                            raise

                print("=== 全処理完了 ===")
                messages.success(request, "更新されました")
                return redirect('daily_report_detail_rol')
            except Exception as e:
                print(f"エラー発生: {str(e)}")
                messages.error(request, f"更新中にエラーが発生しました: {str(e)}")
        else:
            print("フォームバリデーションエラー:")
            print("Form errors:", form.errors)
            print("Time form errors:", time_form.errors)

    else:
        print("=== GET処理開始 ===")
        form = DailyReportForm(instance=report)
        time_form = TimeForm(initial={
            'departure_time': report.departure_time,
            'arrival_time': report.arrival_time,
            'opening_time': report.opening_time,
            'sold_out_time': report.sold_out_time,
            'closing_time': report.closing_time,
        })

    context = {
        'form': form,
        'time_form': time_form,
        'report': report,
        'entries': entries,
    }
    print("=== レンダリング開始 ===")
    return render(request, 'daily_report_edit_rol.html', context)

# 削除ビュー
@login_required
def daily_report_delete(request, pk):
    report = get_object_or_404(DailyReport, pk=pk)
    if request.method == "POST":
        report.delete()
        messages.success(request, '削除しました。')
        return redirect('daily_report_detail', date=report.date)
    return redirect('daily_report_detail', date=report.date)  # GETリクエスト時は詳細ページにリダイレクト

@require_POST
def update_confirmation(request, report_id):
    try:
        report = DailyReport.objects.get(pk=report_id)
        data = json.loads(request.body)
        report.confirmed = data['confirmed']
        report.save()
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error'}, status=400)


# メニューデータ(Product)の一覧表示、編集、削除
@login_required
def product_list_view(request):
    # weekごとに集約して表示
    weeks = Product.objects.order_by('-week').values('week').distinct()
    context = {
        'weeks': weeks,
    }
    return render(request, 'product_list.html', context)

@login_required
def product_detail_view(request, week):
    # 指定されたweekに対応するすべてのProductオブジェクトを取得
    products = Product.objects.filter(week=week)

    if request.method == 'POST':
        # 各フォームを作成し、POST データを処理
        forms = [ProductForm(request.POST, instance=product, prefix=f'form_{i}') for i, product in enumerate(products)]
        if all(form.is_valid() for form in forms):
            for form in forms:
                form.save()
            return redirect('product_list')
    else:
        # 編集フォームの準備
        forms = [ProductForm(instance=product, prefix=f'form_{i}') for i, product in enumerate(products)]

    context = {
        'week': week,
        'products': products,
        'forms': forms,
    }
    return render(request, 'product_detail.html', context)
@login_required
def delete_product_view(request, week):
    # 指定されたweekに対応するすべてのProductオブジェクトを削除
    Product.objects.filter(week=week).delete()
    return redirect('product_list')

@login_required
def location_list_view(request):
    if request.method == 'POST':
        # 削除リクエストの場合
        if 'delete' in request.POST:
            selected_ids = request.POST.getlist('selected_ids')
            SalesLocation.objects.filter(id__in=selected_ids).delete()
            messages.success(request, '選択した販売場所データを削除しました。')
            return HttpResponseRedirect(reverse('location_list'))

        # 上書き保存のリクエスト
        location_data = request.POST
        for i in range(len(location_data)//7):  # 各locationごとのキー数が7つのため調整
            no = location_data.get(f'location[{i}][no]')
            name = location_data.get(f'location[{i}][name]')
            loc_type = location_data.get(f'location[{i}][type]')
            price_type = location_data.get(f'location[{i}][price_type]')
            service_name = location_data.get(f'location[{i}][service_name]')
            service_price = location_data.get(f'location[{i}][service_price]')
            direct_return = location_data.get(f'location[{i}][direct_return]')

            # データを保存
            SalesLocation.objects.update_or_create(
                no=no,
                defaults={
                    'name': name,
                    'type': loc_type,
                    'price_type': price_type,
                    'service_name': service_name,
                    'service_price': service_price,
                    'direct_return': direct_return
                }
            )

        messages.success(request, '販売場所データを更新しました。')
        return HttpResponseRedirect(reverse('location_list'))

    # データを取得して表示
    locations = SalesLocation.objects.all().order_by('no')
    return render(request, 'location_list.html', {'locations': locations})

@login_required
def others_list_view(request):
    if request.method == 'POST':
    # 削除リクエストの場合
        if 'delete' in request.POST:
            selected_ids = request.POST.getlist('selected_ids')
            OthersItem.objects.filter(id__in=selected_ids).delete()
            messages.success(request, '選択した項目を削除しました。')
            return HttpResponseRedirect(reverse('others_item_list'))

        # 上書き保存のリクエスト
        others_data = request.POST
        for i in range(len(others_data)//3):  # 各locationごとのキー数が7つのため調整
            no = others_data.get(f'others[{i}][no]')
            name = others_data.get(f'others[{i}][name]')
            price = others_data.get(f'others[{i}][price]')


        # データを保存
        OthersItem.objects.update_or_create(
            no=no,
            defaults={
                'no': no,
                'name': name,
                'price': price,
            }
        )

        messages.success(request, 'その他の項目を更新しました。')
        return HttpResponseRedirect(reverse('others_item_list'))

    # データを取得して表示
    item = OthersItem.objects.all().order_by('no')
    return render(request, 'others_item_list.html', {'item':item})




# 持参数データ(ItemQuantity)の一覧表示、編集、削除
@login_required
def item_quantity_list_view(request):

    item_dates = ItemQuantity.objects.order_by('-target_date').values('target_date','target_week').distinct()  # target_date のみを取得
    item_date = []
    for item in item_dates:
        item_date.append({
            'target_date': datetime.strptime(item['target_date'], '%Y-%m-%d') , # 適切なフォーマットを指定
            'target_week': datetime.strptime(item['target_week'], '%Y-%m-%d')
        })
    context = {
        'item_date': item_date
    }
    return render(request, 'item_quantity_list.html', context)

@login_required
def item_quantity_detail_view(request, date):
    # date_strをそのまま使用（str型）
    # 日付の形式を確認する（オプション）
    # date_strの形式をそのまま使用して、表示用や他の処理に使用可能
    # ここでは単純にdate_strを表示に使ったり、他の処理を行う
    date_obj = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')  # フォーマットに合わせて変換
    date_only = date_obj.date()  # datetimeオブジェクトからdateオブジェクトを取得
    date_str = date_only.strftime('%Y-%m-%d')  # dateオブジェクトを文字列に変換
    print(f"Selected date: {date_str}")  # 確認用

    # 商品ごとの数量を取得
    item_quantities = ItemQuantity.objects.filter(target_date=date_str) \
                                          .values('product__name') \
                                          .annotate(total_quantity=Sum('quantity')) \
                                          .order_by('product__no')

    # 各商品ごとに、販売場所ごとの数量を取得
    item_details = {}
    sales_locations = []  # 販売場所を格納するリスト

    # 合計を格納する辞書
    location_totals = {}

    # すべての販売場所を取得し、番号でソート
    all_sales_locations = SalesLocation.objects.order_by('no')  # sales_location_noでソート

    for item in item_quantities:
        product_name = item['product__name']
        # 販売場所別の数量を集約
        sales_location_quantities = ItemQuantity.objects.filter(
            product__name=product_name,
            target_date=date_str
        ).values('sales_location__name', 'sales_location__no').annotate(
            location_quantity=Sum('quantity')
        ).order_by('sales_location__no')

        # 各商品ごとのデータを準備
        item_details[product_name] = {location['sales_location__name']: location['location_quantity'] for location in sales_location_quantities}

        # 販売場所をセットに追加し、合計を計算
        for location in sales_location_quantities:
            location_name = location['sales_location__name']
            quantity = location['location_quantity']
            if location_name not in location_totals:
                location_totals[location_name] = 0
            location_totals[location_name] += quantity

    # sales_locationsをall_sales_locationsに基づいてリストに整形
    sales_locations = [location.name for location in all_sales_locations]

    # テンプレートに渡すデータを準備
    context = {
        'item_quantities': item_quantities,  # 集計した商品の数量データ
        'item_details': item_details,  # 商品ごとに販売場所別の数量を集約
        'sales_locations': sales_locations,  # 販売場所名のリスト
        'selected_date': date_str,  # 選択された日付
        'location_totals': location_totals,  # 販売場所ごとの合計
    }

    return render(request, 'item_quantity_detail.html', context)


@login_required
def item_quantity_update_view(request, pk):
    item_quantity = get_object_or_404(ItemQuantity, pk=pk)
    if request.method == 'POST':
        form = ItemQuantityForm(request.POST, instance=item_quantity)
        if form.is_valid():
            form.save()
            return redirect('item_quantity_list')
    else:
        form = ItemQuantityForm(instance=item_quantity)
    return render(request, 'item_quantity_form.html', {'form': form})



# 実績データ確認ビュー
@login_required
def performance_data_view(request):
    # 現在の日付
    today = datetime.today()
    current_year = today.year
    current_month = today.month

    # GETリクエストから日付範囲を取得
    search_date_start = request.GET.get('search_date_start')
    search_date_end = request.GET.get('search_date_end')
    date_range_option = request.GET.get('range')

    # ボタンで選択された範囲に基づいて日付を設定
    if date_range_option == 'this_month':
        # 今月の初日から今日までの範囲
        start_date = make_aware(datetime(current_year, current_month, 1))
        end_date = make_aware(today + timedelta(days=1))  # 本日を含めるために1日追加
    elif date_range_option == 'last_month':
        # 先月の初日から月末までの範囲
        last_month = current_month - 1 if current_month > 1 else 12
        last_year = current_year if current_month > 1 else current_year - 1
        start_date = make_aware(datetime(last_year, last_month, 1))
        end_date = make_aware(datetime(last_year, last_month, (today.replace(day=1) - timedelta(days=1)).day))
    elif date_range_option == 'two_months_ago':
        # 先々月の初日から月末までの範囲
        two_months_ago = current_month - 2 if current_month > 2 else 12 + (current_month - 2)
        two_years_ago = current_year if current_month > 2 else current_year - 1
        start_date = make_aware(datetime(two_years_ago, two_months_ago, 1))
        end_date = make_aware(datetime(two_years_ago, two_months_ago, (today.replace(day=1) - timedelta(days=1)).day))
    elif search_date_start and search_date_end:
        # ユーザーが指定した日付範囲
        start_date = make_aware(datetime.strptime(search_date_start, '%Y-%m-%d'))
        end_date = make_aware(datetime.strptime(search_date_end, '%Y-%m-%d')) + timedelta(days=1)
    else:
        # デフォルトは今月
        start_date = make_aware(today.replace(day=1))
        end_date = make_aware(today + timedelta(days=1))  # 今日を含めるために1日追加

    # 日付範囲に基づいてレポートを取得
    reports = DailyReport.objects.filter(date__range=[start_date, end_date])

    # 日付ごとに集計データを取得
    date_range = (end_date - start_date).days
    summary_data = []

    total_quantity = 0
    total_sales_quantity = 0
    total_remaining = 0
    total_revenue = 0
    cash = 0
    paypay = 0
    digital_payment = 0

    for i in range(date_range):
        current_date = start_date + timedelta(days=i)
        daily_reports = reports.filter(date=current_date)

        daily_total_quantity = daily_reports.aggregate(total_quantity=Sum('total_quantity'))['total_quantity'] or 0
        daily_total_sales_quantity = daily_reports.aggregate(total_sales_quantity=Sum('total_sales_quantity'))['total_sales_quantity'] or 0
        daily_total_remaining = daily_reports.aggregate(total_remaining=Sum('total_remaining'))['total_remaining'] or 0
        daily_total_revenue = daily_reports.aggregate(total_revenue=Sum('total_revenue'))['total_revenue'] or 0
        daily_cash = daily_reports.aggregate(cash=Sum('cash'))['cash'] or 0
        daily_paypay = daily_reports.aggregate(paypay=Sum('paypay'))['paypay'] or 0
        daily_digital_payment = daily_reports.aggregate(digital_payment=Sum('digital_payment'))['digital_payment'] or 0

        # 日ごとの廃棄率計算
        waste_rate = 0 if daily_total_quantity == 0 else (daily_total_remaining / daily_total_quantity) * 100

        summary_data.append({
            'date': current_date,
            'total_quantity': daily_total_quantity,
            'total_sales_quantity': daily_total_sales_quantity,
            'total_remaining': daily_total_remaining,
            'total_revenue': daily_total_revenue,
            'cash': daily_cash,
            'paypay': daily_paypay,
            'digital_payment': daily_digital_payment,
            'waste_rate': waste_rate,
        })

        # 合計値の計算
        total_quantity += daily_total_quantity
        total_sales_quantity += daily_total_sales_quantity
        total_remaining += daily_total_remaining
        total_revenue += daily_total_revenue
        cash += daily_cash
        paypay += daily_paypay
        digital_payment += daily_digital_payment

    # 全体の廃棄率計算
    total_waste_rate = 0 if total_quantity == 0 else (total_remaining / total_quantity) * 100

    # テンプレートにデータを渡す
    context = {
        'summary_data': summary_data,
        'search_date_start': search_date_start,
        'search_date_end': search_date_end,
        'total_quantity': total_quantity,
        'total_sales_quantity': total_sales_quantity,
        'total_remaining': total_remaining,
        'total_revenue': total_revenue,
        'cash': cash,
        'paypay': paypay,
        'digital_payment': digital_payment,
        'total_waste_rate': total_waste_rate,
    }

    return render(request, 'performance_data_template.html', context)

@login_required
def menu_sales_performance_view(request, year, month, day):
    selected_date = date(year, month, day)

    summary_data = DailyReportEntry.objects.filter(report__date=selected_date).values(
    'product',
    'product_no',  # productの識別子を使用
    ).annotate(
        total_quantity=Sum('quantity'),
        total_sales_quantity=Sum('sales_quantity'),
        total_remaining=Sum('remaining_number'),
        popular_count=Count('id', filter=Q(popular=True)),
        unpopular_count=Count('id', filter=Q(unpopular=True)),
    ).order_by('product_no')  # no順にソート

    # 合計値を計算
    summary_data_filtered = summary_data.filter(product_no__range=(1, 10))  # product_noが1〜10の範囲に絞り込み
    total_quantity_sum = summary_data_filtered.aggregate(Sum('total_quantity'))['total_quantity__sum'] or 0
    total_sales_quantity_sum = summary_data_filtered.aggregate(Sum('total_sales_quantity'))['total_sales_quantity__sum'] or 0
    total_remaining_sum = summary_data_filtered.aggregate(Sum('total_remaining'))['total_remaining__sum'] or 0
    total_popular_count = summary_data_filtered.aggregate(Sum('popular_count'))['popular_count__sum'] or 0
    total_unpopular_count = summary_data_filtered.aggregate(Sum('unpopular_count'))['unpopular_count__sum'] or 0

    # 廃棄率の計算
    summary_with_waste_rate = []
    for entry in summary_data:
        waste_rate = (entry['total_remaining'] / entry['total_quantity'] * 100) if entry['total_quantity'] > 0 else 0
        entry['waste_rate'] = waste_rate
        summary_with_waste_rate.append(entry)

    # 合計行の廃棄率も計算
    total_waste_rate = (total_remaining_sum / total_quantity_sum * 100) if total_quantity_sum > 0 else 0

    context = {
        'selected_date': selected_date,
        'summary_data': summary_with_waste_rate,
        'total_quantity_sum': total_quantity_sum,
        'total_sales_quantity_sum': total_sales_quantity_sum,
        'total_remaining_sum': total_remaining_sum,
        'total_popular_count': total_popular_count,
        'total_unpopular_count': total_unpopular_count,
        'total_waste_rate': total_waste_rate,
    }
    return render(request, 'menu_sales_performance_template.html', context)




# 販売場所別実績データビュー
@login_required
def performance_by_location_view(request):
    # 現在の年と月を初期値として取得する
    current_year = date.today().year
    current_month = date.today().month
    # 検索された年と月を取得（指定がない場合は現在の年と月を使用）
    search_year = int(request.GET.get('search_year', current_year))
    search_month = int(request.GET.get('search_month', current_month))

    # 販売場所の選択リストを取得
    selected_locations = request.GET.getlist('sales_locations')  # リスト形式で取得

    # 現在の月の全日付を取得
    start_date = date(search_year, search_month, 1)
    end_date = start_date + relativedelta(months=1) - timedelta(days=1)
    days_in_month = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    # 日本語の曜日リスト（インデックス: 0=月曜日, 1=火曜日...6=日曜日）
    JAPANESE_WEEKDAYS = ['月', '火', '水', '木', '金', '土', '日']

    # 各日付に対応する日本語の曜日を取得
    days_with_weekdays = [
        (day, JAPANESE_WEEKDAYS[day.weekday()]) for day in days_in_month
    ]

    # 対象月のデータを取得
    reports = DailyReport.objects.filter(date__range=(start_date, end_date))

    # 選択された販売場所に基づいてフィルタリング
# 「全ての販売場所」もしくは未選択なら全ての販売場所を表示
    if 'all' in selected_locations or not selected_locations:
        locations = SalesLocation.objects.all()
    else:
        # 選択された販売場所だけを表示
        locations = SalesLocation.objects.filter(id__in=selected_locations)


    # 販売場所ごとの日別販売数データを集計
    summary_data = []
    for location in locations:
        day_reports = reports.filter(location=location)

        # 日ごとの合計販売数をリスト化
        sales_per_day = []
        for day in range(1, end_date.day + 1):
            day_date = date(int(search_year), int(search_month), day)
            total_sales = day_reports.filter(date=day_date).aggregate(total_sales=Sum('total_sales_quantity'))['total_sales'] or 0
            sales_per_day.append(total_sales)

        # 0を除外して平均値を計算、整数に丸める
        valid_sales = [sale for sale in sales_per_day if sale > 0]  # 1以上のデータのみを含む
        if valid_sales:
            average_sales = sum(valid_sales) // len(valid_sales)  # 整数で計算
        else:
            average_sales = 0

        # 前月のデータを取得し、前月の平均値を計算（こちらも0を除外）
        previous_month_start = start_date - relativedelta(months=1)
        previous_month_end = previous_month_start + relativedelta(months=1) - timedelta(days=1)
        previous_month_reports = DailyReport.objects.filter(location=location, date__range=(previous_month_start, previous_month_end))

        previous_month_sales = previous_month_reports.values_list('total_sales_quantity', flat=True)
        previous_month_sales = [sale for sale in previous_month_sales if sale > 0]  # 0を除外

        if previous_month_sales:
            previous_month_average = sum(previous_month_sales) // len(previous_month_sales)  # 整数で計算
        else:
            previous_month_average = 0

        # 前月比を整数で計算
        if previous_month_average > 0:
            comparison_to_last_month = average_sales - previous_month_average
            comparison_to_last_month_abs = abs(comparison_to_last_month)  # 絶対値を計算
        else:
            comparison_to_last_month = None
            comparison_to_last_month_abs = None

        # データをリストに追加
        summary_data.append({
            'location': location,
            'sales': sales_per_day,
            'average_sales': average_sales,
            'comparison_to_last_month': comparison_to_last_month,
            'comparison_to_last_month_abs': comparison_to_last_month_abs,  # 絶対値も追加
        })

    context = {
        'summary_data': summary_data,
        'days_in_month': days_with_weekdays,
        'search_year': search_year,  # テンプレートに渡す
        'search_month': search_month,  # テンプレートに渡す
        'locations': SalesLocation.objects.all(),  # 販売場所の全リストを渡す
        'selected_locations': selected_locations,  # 選択された販売場所を渡す
    }

    return render(request, 'performance_by_location_template.html', context)

@login_required
def performance_by_location_view_rol(request):
    # 現在の年と月を初期値として取得する
    current_year = date.today().year
    current_month = date.today().month
    # 検索された年と月を取得（指定がない場合は現在の年と月を使用）
    search_year = int(request.GET.get('search_year', current_year))
    search_month = int(request.GET.get('search_month', current_month))

    # 販売場所の選択リストを取得
    selected_locations = request.GET.getlist('sales_locations')  # リスト形式で取得

    # 現在の月の全日付を取得
    start_date = date(search_year, search_month, 1)
    end_date = start_date + relativedelta(months=1) - timedelta(days=1)
    days_in_month = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    # 日本語の曜日リスト（インデックス: 0=月曜日, 1=火曜日...6=日曜日）
    JAPANESE_WEEKDAYS = ['月', '火', '水', '木', '金', '土', '日']

    # 各日付に対応する日本語の曜日を取得
    days_with_weekdays = [
        (day, JAPANESE_WEEKDAYS[day.weekday()]) for day in days_in_month
    ]

    # 対象月のデータを取得
    reports = DailyReport.objects.filter(date__range=(start_date, end_date))

    # 選択された販売場所に基づいてフィルタリング
# 「全ての販売場所」もしくは未選択なら全ての販売場所を表示
    if 'all' in selected_locations or not selected_locations:
        locations = SalesLocation.objects.all()
    else:
        # 選択された販売場所だけを表示
        locations = SalesLocation.objects.filter(id__in=selected_locations)


    # 販売場所ごとの日別販売数データを集計
    summary_data = []
    for location in locations:
        day_reports = reports.filter(location=location)

        # 日ごとの合計販売数をリスト化
        sales_per_day = []
        for day in range(1, end_date.day + 1):
            day_date = date(int(search_year), int(search_month), day)
            total_sales = day_reports.filter(date=day_date).aggregate(total_sales=Sum('total_sales_quantity'))['total_sales'] or 0
            sales_per_day.append(total_sales)

        # 0を除外して平均値を計算、整数に丸める
        valid_sales = [sale for sale in sales_per_day if sale > 0]  # 1以上のデータのみを含む
        if valid_sales:
            average_sales = sum(valid_sales) // len(valid_sales)  # 整数で計算
        else:
            average_sales = 0

        # 前月のデータを取得し、前月の平均値を計算（こちらも0を除外）
        previous_month_start = start_date - relativedelta(months=1)
        previous_month_end = previous_month_start + relativedelta(months=1) - timedelta(days=1)
        previous_month_reports = DailyReport.objects.filter(location=location, date__range=(previous_month_start, previous_month_end))

        previous_month_sales = previous_month_reports.values_list('total_sales_quantity', flat=True)
        previous_month_sales = [sale for sale in previous_month_sales if sale > 0]  # 0を除外

        if previous_month_sales:
            previous_month_average = sum(previous_month_sales) // len(previous_month_sales)  # 整数で計算
        else:
            previous_month_average = 0

        # 前月比を整数で計算
        if previous_month_average > 0:
            comparison_to_last_month = average_sales - previous_month_average
            comparison_to_last_month_abs = abs(comparison_to_last_month)  # 絶対値を計算
        else:
            comparison_to_last_month = None
            comparison_to_last_month_abs = None

        # データをリストに追加
        summary_data.append({
            'location': location,
            'sales': sales_per_day,
            'average_sales': average_sales,
            'comparison_to_last_month': comparison_to_last_month,
            'comparison_to_last_month_abs': comparison_to_last_month_abs,  # 絶対値も追加
        })

    context = {
        'summary_data': summary_data,
        'days_in_month': days_with_weekdays,
        'search_year': search_year,  # テンプレートに渡す
        'search_month': search_month,  # テンプレートに渡す
        'locations': SalesLocation.objects.all(),  # 販売場所の全リストを渡す
        'selected_locations': selected_locations,  # 選択された販売場所を渡す
    }

    return render(request, 'performance_by_location_template_rol.html', context)

# CSVダウンロード機能
@login_required
def download_csv(request):
    # GETリクエストから日付範囲を取得
    search_date_start = request.GET.get('search_date_start')
    search_date_end = request.GET.get('search_date_end')

    # デフォルト値として今日の日付を設定
    today = datetime.today().date()

    try:
        if search_date_start and search_date_end:
            # 文字列から日付オブジェクトに変換
            start_date = datetime.strptime(search_date_start, '%Y-%m-%d').date()
            end_date = datetime.strptime(search_date_end, '%Y-%m-%d').date()
            end_date = end_date + timedelta(days=1)
        else:
            # デフォルトは今日の日付を範囲とする
            start_date = today
            end_date = today + timedelta(days=1)
    except ValueError:
        # 日付の変換に失敗した場合は今日の日付を使用
        start_date = today
        end_date = today + timedelta(days=1)

    # 日付範囲に基づいてレポートを取得
    reports = DailyReport.objects.filter(date__range=[start_date, end_date])

    # レスポンスをCSV形式で作成（Shift-JISエンコーディングを指定）
    response = HttpResponse(content_type='text/csv; charset=shift-jis')
    filename = "販売実績集計データ.csv"
    filename_encoded = urllib.parse.quote(filename)
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename_encoded}'

    # Shift-JIS用のCSVライターを作成
    writer = csv.writer(response)
    
    # ヘッダーをShift-JISでエンコード
    header = ['日付', '総持参数', '総販売数', '総残数', '売上総額(¥)', '現金(¥)', 'PayPay(¥)', '電子決済(¥)', '廃棄率(%)']
    writer.writerow([h.encode('shift-jis', 'ignore').decode('shift-jis') for h in header])

    # 日付ごとに集計データを取得してCSVに書き込む
    date_range = (end_date - start_date).days

    for i in range(date_range):
        current_date = start_date + timedelta(days=i)
        daily_reports = reports.filter(date=current_date)

        total_quantity = daily_reports.aggregate(total_quantity=Sum('total_quantity'))['total_quantity'] or 0
        total_sales_quantity = daily_reports.aggregate(total_sales_quantity=Sum('total_sales_quantity'))['total_sales_quantity'] or 0
        total_remaining = daily_reports.aggregate(total_remaining=Sum('total_remaining'))['total_remaining'] or 0
        total_revenue = daily_reports.aggregate(total_revenue=Sum('total_revenue'))['total_revenue'] or 0
        cash = daily_reports.aggregate(cash=Sum('cash'))['cash'] or 0
        paypay = daily_reports.aggregate(paypay=Sum('paypay'))['paypay'] or 0
        digital_payment = daily_reports.aggregate(digital_payment=Sum('digital_payment'))['digital_payment'] or 0
        waste_rate = 0 if total_quantity == 0 else (total_remaining / total_quantity) * 100

        # 日付を文字列に変換
        date_str = current_date.strftime('%Y-%m-%d')
        
        # データ行を作成してShift-JISでエンコード
        row = [
            date_str,
            str(total_quantity),
            str(total_sales_quantity),
            str(total_remaining),
            str(total_revenue),
            str(cash),
            str(paypay),
            str(digital_payment),
            f"{waste_rate:.1f}"
        ]
        writer.writerow([str(item).encode('shift-jis', 'ignore').decode('shift-jis') for item in row])

    return response

# CSVダウンロード機能(allreport)
@login_required
def download_csv_allreport(request):
    # HTTPレスポンスにCSVのヘッダーを設定（文字コードをShift-JISに指定）
    response = HttpResponse(content_type='text/csv; charset=shift-jis')
    filename = "日計表送信データ.csv"
    filename_encoded = urllib.parse.quote(filename)
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename_encoded}'

    # Shift-JIS用のCSVライターを作成
    writer = csv.writer(response)

    # モデルの全フィールドを取得
    fields = [field.name for field in DailyReport._meta.get_fields()]

    # CSVヘッダーを書き込み（Shift-JISでエンコード）
    header = ['レポートID', '販売場所NO', '日付', '販売場所', '担当者', '天気', '気温', '総持参数', 
              '単価別販売数1', '単価別販売数2', '単価別販売数3', '総販売数', '総残数', 
              'その他1項目', 'その他1単価', 'その他1販売数', 'その他2項目', 'その他2単価', 'その他2販売数', 
              'その他売上合計', '総売上', 'ご飯なし', 'ご飯追加', 'クーポン600', 'クーポン700', 
              '割引・返金50円', '割引・返金100円', 'サービス名', 'サービス単価', 'サービス600', 
              'サービス700', 'サービス100', '割引合計', 'PayPay', '電子決済', '現金', '差額', 
              '出発時間', '到着時間', '開店時間', '完売時間', '閉店時間', 'ガソリン代', '高速代', 
              '駐車場代', 'パート代', 'その他経費', 'コメント', '明日の食数設定', '確認済み', '更新日時',
              '商品名1', '商品NO1', '持参数1', '販売数1', '残数1', '売上1', '完売1', '人気1', '不人気1',
              '商品名2', '商品NO2', '持参数2', '販売数2', '残数2', '売上2', '完売2', '人気2', '不人気2',
              '商品名3', '商品NO3', '持参数3', '販売数3', '残数3', '売上3', '完売3', '人気3', '不人気3',
              '商品名4', '商品NO4', '持参数4', '販売数4', '残数4', '売上4', '完売4', '人気4', '不人気4',
              '商品名5', '商品NO5', '持参数5', '販売数5', '残数5', '売上5', '完売5', '人気5', '不人気5',
              '商品名6', '商品NO6', '持参数6', '販売数6', '残数6', '売上6', '完売6', '人気6', '不人気6',
              '商品名7', '商品NO7', '持参数7', '販売数7', '残数7', '売上7', '完売7', '人気7', '不人気7',
              '商品名8', '商品NO8', '持参数8', '販売数8', '残数8', '売上8', '完売8', '人気8', '不人気8',
              '商品名9', '商品NO9', '持参数9', '販売数9', '残数9', '売上9', '完売9', '人気9', '不人気9',
              '商品名10', '商品NO10', '持参数10', '販売数10', '残数10', '売上10', '完売10', '人気10', '不人気10',
              '商品名11', '商品NO11', '持参数11', '販売数11', '残数11', '売上11', '完売11',
              ]# ヘッダーにエントリのフィールドを追加

    # 文字列をShift-JISでエンコード
    encoded_header = [h.encode('shift-jis', 'ignore').decode('shift-jis') for h in header]
    writer.writerow(encoded_header)

    # 日付範囲で絞り込み
    search_date_start = request.GET.get('search_date_start')
    search_date_end = request.GET.get('search_date_end')

    if search_date_start and search_date_end:
        start_date = datetime.strptime(search_date_start, '%Y-%m-%d')
        end_date = datetime.strptime(search_date_end, '%Y-%m-%d')
        reports = DailyReport.objects.filter(date__range=(start_date, end_date))
    else:
        reports = DailyReport.objects.all()

    # レコードを書き込み（文字列をShift-JISでエンコード）
    for report in reports:
        # DailyReport のデータを行に追加
        row = [
            report.id,  # レポートID
            report.location_no,  # 販売場所NO
            report.date,  # 日付
            report.location,  # 販売場所
            report.person_in_charge,  # 担当者
            report.weather,  # 天気
            report.temp,  # 気温
            report.total_quantity,  # 持参数合計
            report.sales_price_quantity_1,  # 単価別販売数1
            report.sales_price_quantity_2,  # 単価別販売数2
            report.sales_price_quantity_3,  # 単価別販売数3
            report.total_sales_quantity,  # 販売数合計
            report.total_remaining,  # 残数合計
            report.others_sales_1,  # その他1項目
            report.others_price1,  # その他1単価
            report.others_sales_quantity1,  # その他1販売数
            report.others_sales_2,  # その他2項目
            report.others_price2,  # その他2単価
            report.others_sales_quantity2,  # その他2販売数
            report.total_others_sales,  # その他売上合計
            report.total_revenue,  # 総売上
            report.no_rice_quantity,  # ご飯なし
            report.extra_rice_quantity,  # ご飯追加
            report.coupon_type_600,  # クーポン600
            report.coupon_type_700,  # クーポン700
            report.discount_50,  # 割引・返金50
            report.discount_100,  # 割引・返金100
            report.service_name,  # サービス名
            report.service_price,  # サービス価格
            report.service_type_600,  # サービス600
            report.service_type_700,  # サービス700
            report.service_type_100,  # サービス100
            report.total_discount,  # 割引合計
            report.paypay,  # PayPay
            report.digital_payment,  # 電子決済
            report.cash,  # 現金
            report.sales_difference,  # 差額
            report.departure_time,  # 出発時間
            report.arrival_time,  # 到着時間
            report.opening_time,  # 開店時間
            report.sold_out_time,  # 完売時間
            report.closing_time,  # 閉店時間
            report.gasolin,  # ガソリン代
            report.highway,  # 高速代
            report.parking,  # 駐車場代
            report.part,  # パート代
            report.others,  # その他経費
            report.comments,  # コメント
            report.food_count_setting,  # 明日の食数設定
            report.confirmed,  # 確認済みフラグ
            report.updated_at  # 更新日時
        ]

        # DailyReportEntry のデータを取得して行に追加
        entries = report.entries.all()
        entry_data = []
        for entry in entries:
            entry_data.extend([
                entry.product_no, entry.product, entry.quantity, entry.sales_quantity,
                entry.remaining_number, entry.total_sales, entry.sold_out, entry.popular, entry.unpopular
            ])
        
        # 1行にまとめて書き込む
        writer.writerow(row + entry_data)

    return response


@login_required
def location_performance_view(request, location_id, search_year, search_month):
    # 対象の販売場所を取得
    location = get_object_or_404(SalesLocation, id=location_id)

    # 対象月の最初と最後の日付を取得
    start_date = date(search_year, search_month, 1)
    end_date = start_date + relativedelta(months=1) - timedelta(days=1)

    # 現在の月の全日付を取得
    days_in_month = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    # 対象月の日ごとの販売データを取得
    reports = DailyReport.objects.filter(location=location, date__range=(start_date, end_date))

    # 曜日を日本語に変換するマッピング
    day_of_week_japanese = ['日', '月', '火', '水', '木', '金', '土']

    # 曜日のラベルを作成（%Y-%m-%dに日本語の曜日を追加）
    labels = [day.strftime('%d') + f" ({day_of_week_japanese[day.weekday()]})" for day in days_in_month]


    # 日ごとの合計販売数をリスト化
    sales_per_day = []
    for day in days_in_month:
        total_sales = reports.filter(date=day).aggregate(total_sales=Sum('total_sales_quantity'))['total_sales'] or 0
        sales_per_day.append(total_sales)

    # 前月のデータを取得
    prev_month_start = start_date - relativedelta(months=1)
    prev_month_end = prev_month_start + relativedelta(months=1) - timedelta(days=1)

    # 前月の販売データを取得（0を除く）
    prev_month_reports = DailyReport.objects.filter(
        location=location, date__range=(prev_month_start, prev_month_end)
    ).exclude(total_sales_quantity=0)

    # 前月の平均販売数を計算（0を除いたデータのみで計算）
    total_sales_prev_month = prev_month_reports.aggregate(total_sales=Sum('total_sales_quantity'))['total_sales'] or 0
    days_in_prev_month = prev_month_reports.count()  # 0を除いた販売がある日の数を使用
    avg_sales_prev_month = total_sales_prev_month / days_in_prev_month if days_in_prev_month > 0 else 0

    # 日付と販売数をペアにする
    days_and_sales = zip(days_in_month, sales_per_day)

    # 戻り先のURLを構築
    previous_url = reverse('location_performance', kwargs={'location_id': location_id, 'search_year': search_year, 'search_month': search_month})

    # コンテキストにデータを渡す
    context = {
        'location': location,
        'search_year': search_year,
        'search_month': search_month,
        'days_in_month': days_in_month,  # 日付リストを直接渡す
        'sales_per_day': sales_per_day,  # 販売数リストを直接渡す
        'avg_sales_prev_month': avg_sales_prev_month,  # 前月の平均販売数（0を除いた）
        'labels': labels,  # 曜日ラベルを追加
        'previous_url': previous_url,  # 戻り先URLを渡す
        'selected_location': location_id,  # 戻るボタン用にlocation_idを追加
    }

    return render(request, 'location_performance_template.html', context)

@login_required
def menu_history_view(request):
    if request.method == "POST":
        menu_name = request.POST.get('menu_name', '').strip()  # メニュー名を取得し、前後の空白を削除

        if menu_name:  # メニュー名が空でない場合
            # メニュー名でフィルタリング
            daily_reports = DailyReportEntry.objects.filter(product__icontains=menu_name).select_related('report')

            # 日付ごとに集計
            summary_data = daily_reports.values('report__date', 'product').annotate(
                total_quantity=Sum('quantity'),
                total_sales_quantity=Sum('sales_quantity'),
                total_remaining=Sum('remaining_number'),
            ).order_by('report__date')

            # 廃棄率を計算
            for entry in summary_data:
                if entry['total_quantity'] > 0:  # 0で割るのを避ける
                    entry['waste_rate'] = (entry['total_remaining'] / entry['total_quantity']) * 100
                else:
                    entry['waste_rate'] = 0  # 持参数が0の場合は廃棄率も0

            return render(request, 'menu_history_template.html', {'summary_data': summary_data, 'menu_name': menu_name})
        else:
            # メニュー名がブランクの場合は何も表示しない
            return render(request, 'menu_history_template.html', {'summary_data': [], 'menu_name': ''})

    return render(request, 'menu_history_template.html')

@login_required
def performance_by_location_calender_view(request, location_id, search_year, search_month):
    # 対象の販売場所を取得
    location = get_object_or_404(SalesLocation, id=location_id)

    # 年と月の選択を処理
    if request.method == "GET":
        search_year = request.GET.get('year', search_year)
        search_month = request.GET.get('month', search_month)

    # 対象月の最初と最後の日付を取得
    start_date = date(int(search_year), int(search_month), 1)
    end_date = start_date + relativedelta(months=1) - timedelta(days=1)

    # 日曜日を最初の曜日として設定
    calendar.setfirstweekday(calendar.SUNDAY)

    # 現在の月の全日付を取得
    days_in_month = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    # カレンダーの生成
    month_days = calendar.monthcalendar(int(search_year), int(search_month))

    # 対象月の日ごとの販売データを取得
    reports = DailyReport.objects.filter(location=location, date__range=(start_date, end_date))

    # 各日付に対する販売データを取得
    performance_data = defaultdict(lambda: {'sales': 0, 'total_quantity': 0, 'weather': None, 'closing_time': None})
    for report in reports:
        day = report.date.day
        performance_data[day]['sales'] += report.total_sales_quantity
        performance_data[day]['total_quantity'] += report.total_quantity  # total_quantityを追加
        if not performance_data[day]['closing_time'] or report.closing_time > performance_data[day]['closing_time']:
            performance_data[day]['closing_time'] = report.closing_time
        if performance_data[day]['weather'] is None:  # weatherを初期化
            performance_data[day]['weather'] = report.weather

    # コンテキストにデータを渡す
    context = {
        'location': location,
        'search_year': search_year,
        'search_month': search_month,
        'performance_data': performance_data,  # 日付ごとの販売データと閉店時間を含む
        'month_days': month_days,  # カレンダーの日付を追加
    }

    # 例: 年と月のリストを作成
    available_years = range(2020, datetime.today().year + 1)  # 2020年から現在の年まで
    available_months = range(1, 13)  # 1月から12月まで

    context.update({
        'available_years': available_years,  # 年のリストを追加
        'available_months': available_months,  # 月のリストを追加
    })

    return render(request, 'performance_by_location_calender.html', context)